import os
from pathlib import Path
from typing import List, Tuple

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer
from reportlab.lib import colors


def _excel_color_to_hex(cell) -> str:
    try:
        fill = getattr(cell, "fill", None)
        if not fill:
            return ""
        fg = getattr(fill, "fgColor", None)
        if not fg:
            return ""
        # openpyxl may provide rgb like 'FFRRGGBB'
        rgb = getattr(fg, "rgb", None)
        if isinstance(rgb, str) and len(rgb) in (6, 8):
            hexrgb = rgb[-6:]
            return f"#{hexrgb}"
        return ""
    except Exception:
        return ""


def _excel_alignment(cell) -> Tuple[str, str]:
    try:
        al = getattr(cell, "alignment", None)
        h = (getattr(al, "horizontal", None) or "left").lower()
        v = (getattr(al, "vertical", None) or "top").lower()
        h_map = {"center": "CENTER", "right": "RIGHT", "left": "LEFT", "justify": "LEFT"}
        v_map = {"center": "MIDDLE", "middle": "MIDDLE", "top": "TOP", "bottom": "BOTTOM"}
        return h_map.get(h, "LEFT"), v_map.get(v, "TOP")
    except Exception:
        return "LEFT", "TOP"


def _compute_dimensions(ws) -> Tuple[int, int, int, int]:
    try:
        dim = ws.calculate_dimension()  # e.g., 'A1:F42'
        if ":" in dim:
            start, end = dim.split(":", 1)
        else:
            start, end = dim, dim
        import openpyxl.utils as xlutils
        min_col, min_row = xlutils.column_index_from_string(''.join([c for c in start if c.isalpha()])), int(''.join([c for c in start if c.isdigit()]))
        max_col, max_row = xlutils.column_index_from_string(''.join([c for c in end if c.isalpha()])), int(''.join([c for c in end if c.isdigit()]))
        # Sanity
        min_col = max(1, min_col)
        min_row = max(1, min_row)
        max_col = max(min_col, max_col)
        max_row = max(min_row, max_row)
        return min_col, min_row, max_col, max_row
    except Exception:
        return 1, 1, ws.max_column or 10, ws.max_row or 50


def convert(xlsx_path: str, pdf_path: str) -> str:
    """
    Convierte un archivo XLSX a PDF en modo apaisado, respetando:
    - fusiones de celdas,
    - colores de fondo básicos,
    - negritas,
    - alineaciones,
    - anchos de columna y alturas de fila aproximados.
    """
    xlsx = Path(xlsx_path)
    out_pdf = Path(pdf_path)
    if not xlsx.exists():
        raise FileNotFoundError(f"XLSX no encontrado: {xlsx}")
    out_pdf.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(str(xlsx), data_only=True)
    story = []

    # Documento PDF
    doc = SimpleDocTemplate(
        str(out_pdf.resolve()),
        pagesize=landscape(A4),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
    )

    # Procesar cada hoja como una tabla
    for ws in wb.worksheets:
        min_c, min_r, max_c, max_r = _compute_dimensions(ws)
        # Construir matriz de datos
        data: List[List[str]] = []
        for r in range(min_r, max_r + 1):
            row_vals: List[str] = []
            for c in range(min_c, max_c + 1):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell):
                    # Relleno vacío en celdas dentro de merge, contenido vive en la esquina sup izq
                    row_vals.append("")
                else:
                    val = cell.value
                    if val is None:
                        txt = ""
                    else:
                        try:
                            txt = str(val)
                        except Exception:
                            txt = ""
                    row_vals.append(txt)
            data.append(row_vals)

        # Anchos de columna (aproximados)
        col_widths: List[float] = []
        for c in range(min_c, max_c + 1):
            cd = ws.column_dimensions.get(get_column_letter(c))
            w = getattr(cd, "width", None)
            # Aproximación: 6.5 puntos por unidad excel
            width_pts = float(w or 12.0) * 6.5
            # mínimo razonable
            col_widths.append(max(24.0, width_pts))

        # Alturas de fila (si disponible)
        row_heights: List[float] = []
        for r in range(min_r, max_r + 1):
            rd = ws.row_dimensions.get(r)
            h = getattr(rd, "height", None)
            # openpyxl height suele estar en puntos
            height_pts = float(h or 18.0)
            row_heights.append(max(14.0, height_pts))

        tbl = Table(data, colWidths=col_widths, rowHeights=row_heights)
        style = TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#888888")),
        ])

        # Fusiones
        try:
            for cr in getattr(ws, "merged_cells", []).ranges:
                r0, r1 = cr.min_row - min_r, cr.max_row - min_r
                c0, c1 = cr.min_col - min_c, cr.max_col - min_c
                style.add("SPAN", (c0, r0), (c1, r1))
        except Exception:
            pass

        # Colores de fondo y alineaciones/negritas por celda
        try:
            for r in range(min_r, max_r + 1):
                for c in range(min_c, max_c + 1):
                    cell = ws.cell(row=r, column=c)
                    # Fondo
                    hexcol = _excel_color_to_hex(cell)
                    if hexcol:
                        style.add("BACKGROUND", (c - min_c, r - min_r), (c - min_c, r - min_r), colors.HexColor(hexcol))
                    # Alineación
                    ha, va = _excel_alignment(cell)
                    style.add("ALIGN", (c - min_c, r - min_r), (c - min_c, r - min_r), ha)
                    style.add("VALIGN", (c - min_c, r - min_r), (c - min_c, r - min_r), va)
                    # Negrita
                    try:
                        if getattr(getattr(cell, "font", None), "bold", False):
                            style.add("FONT", (c - min_c, r - min_r), (c - min_c, r - min_r), "Helvetica-Bold")
                    except Exception:
                        pass
        except Exception:
            pass

        tbl.setStyle(style)
        story.append(tbl)
        story.append(Spacer(1, 12))

    # Generar PDF
    doc.build(story)
    wb.close()
    return str(out_pdf)