import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os
import re

def copy_exact_structure(source_ws, target_ws, start_row=1, target_start_row=1, num_rows=None):
    """Copia la estructura exacta, formato y dimensiones de un rango de celdas"""
    
    if num_rows is None:
        num_rows = source_ws.max_row
    
    # Copiar dimensiones de columnas (solo una vez)
    if target_start_row == 1:
        for col in range(1, source_ws.max_column + 1):
            col_letter = get_column_letter(col)
            source_width = source_ws.column_dimensions[col_letter].width
            if source_width:
                target_ws.column_dimensions[col_letter].width = source_width
    
    # Copiar cada celda con su formato completo
    for row_offset in range(num_rows):
        source_row = start_row + row_offset
        target_row = target_start_row + row_offset
        
        # Copiar altura de fila
        source_height = source_ws.row_dimensions[source_row].height
        if source_height:
            target_ws.row_dimensions[target_row].height = source_height
        
        for col in range(1, source_ws.max_column + 1):
            source_cell = source_ws.cell(row=source_row, column=col)
            target_cell = target_ws.cell(row=target_row, column=col)
            
            # Copiar valor
            target_cell.value = source_cell.value
            
            # Copiar formato de fuente
            if source_cell.font:
                target_cell.font = Font(
                    name=source_cell.font.name,
                    size=source_cell.font.size,
                    bold=source_cell.font.bold,
                    italic=source_cell.font.italic,
                    color=source_cell.font.color
                )
            
            # Copiar alineación
            if source_cell.alignment:
                target_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical,
                    wrap_text=source_cell.alignment.wrap_text,
                    text_rotation=source_cell.alignment.text_rotation,
                    indent=source_cell.alignment.indent
                )
            
            # Copiar relleno
            if source_cell.fill:
                target_cell.fill = PatternFill(
                    start_color=source_cell.fill.start_color,
                    end_color=source_cell.fill.end_color,
                    fill_type=source_cell.fill.fill_type
                )
            
            # Copiar bordes
            if source_cell.border:
                target_cell.border = Border(
                    left=Side(style=source_cell.border.left.style, color=source_cell.border.left.color),
                    right=Side(style=source_cell.border.right.style, color=source_cell.border.right.color),
                    top=Side(style=source_cell.border.top.style, color=source_cell.border.top.color),
                    bottom=Side(style=source_cell.border.bottom.style, color=source_cell.border.bottom.color)
                )


def copy_format_only(source_ws, target_ws, start_row=1, target_start_row=1, num_rows=None):
    """Copia SOLO el formato sin los valores para evitar duplicaciones"""
    
    if num_rows is None:
        num_rows = source_ws.max_row
    
    # Copiar dimensiones de columnas (solo una vez)
    if target_start_row == 1:
        for col in range(1, source_ws.max_column + 1):
            col_letter = get_column_letter(col)
            source_width = source_ws.column_dimensions[col_letter].width
            if source_width:
                target_ws.column_dimensions[col_letter].width = source_width
    
    # Copiar cada celda con su formato completo PERO SIN VALOR
    for row_offset in range(num_rows):
        source_row = start_row + row_offset
        target_row = target_start_row + row_offset
        
        # Copiar altura de fila
        source_height = source_ws.row_dimensions[source_row].height
        if source_height:
            target_ws.row_dimensions[target_row].height = source_height
        
        for col in range(1, source_ws.max_column + 1):
            source_cell = source_ws.cell(row=source_row, column=col)
            target_cell = target_ws.cell(row=target_row, column=col)
            
            # NO copiar valor - esto es clave para evitar duplicaciones
            
            # Copiar formato de fuente
            if source_cell.font:
                target_cell.font = Font(
                    name=source_cell.font.name,
                    size=source_cell.font.size,
                    bold=source_cell.font.bold,
                    italic=source_cell.font.italic,
                    color=source_cell.font.color
                )
            
            # Copiar alineación
            if source_cell.alignment:
                target_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical,
                    wrap_text=source_cell.alignment.wrap_text,
                    text_rotation=source_cell.alignment.text_rotation,
                    indent=source_cell.alignment.indent
                )
            
            # Copiar relleno
            if source_cell.fill:
                target_cell.fill = PatternFill(
                    start_color=source_cell.fill.start_color,
                    end_color=source_cell.fill.end_color,
                    fill_type=source_cell.fill.fill_type
                )
            
            # Copiar bordes
            if source_cell.border:
                target_cell.border = Border(
                    left=Side(style=source_cell.border.left.style, color=source_cell.border.left.color),
                    right=Side(style=source_cell.border.right.style, color=source_cell.border.right.color),
                    top=Side(style=source_cell.border.top.style, color=source_cell.border.top.color),
                    bottom=Side(style=source_cell.border.bottom.style, color=source_cell.border.bottom.color)
                )

def copy_merged_cells(source_ws, target_ws, row_offset=0):
    """Copia las celdas combinadas ajustando las filas según el offset"""
    for merged_range in source_ws.merged_cells.ranges:
        # Ajustar el rango según el offset de filas
        min_row = merged_range.min_row + row_offset
        max_row = merged_range.max_row + row_offset
        min_col = merged_range.min_col
        max_col = merged_range.max_col
        
        # Crear el nuevo rango
        new_range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        target_ws.merge_cells(new_range)

def adapt_jinja_variables_for_day(cell_value, day_num):
    """Adapta las variables Jinja2 para un día específico evitando duplicaciones"""
    if not cell_value or not isinstance(cell_value, str):
        return cell_value
    
    # Si la celda ya tiene variables específicas de día, no la modifiques
    if f"dia_{day_num}" in cell_value or f"ejercicio_dia_{day_num}" in cell_value:
        return cell_value
    
    # Si la celda tiene variables de otros días, no la modifiques
    for other_day in range(1, 6):  # Máximo 5 días
        if other_day != day_num and (f"dia_{other_day}" in cell_value or f"ejercicio_dia_{other_day}" in cell_value):
            return cell_value
    
    # Mapeo de variables originales a variables por día (solo variables genéricas)
    variable_mapping = {
        "{{dia}}": f"{{{{ dia_{day_num} }}}}",
        "{{Ejercicio}}": f"{{{{ ejercicio_dia_{day_num} }}}}",
        "{{Series ejercicio}}": f"{{{{ series_dia_{day_num} }}}}",
        "{{Repeticiones ejercicio}}": f"{{{{ repeticiones_dia_{day_num} }}}}"
    }
    
    adapted_value = cell_value
    for original, adapted in variable_mapping.items():
        # Solo reemplazar si la variable original existe exactamente
        if original in adapted_value:
            adapted_value = adapted_value.replace(original, adapted)
    
    return adapted_value

def detect_week_columns(ws, scan_start_row: int = 1, scan_end_row: int = None):
    """Detecta columnas que representan semanas (S1/S2/S3/S4 o 'Semana 1', 'W1').
    Escanea un rango de filas y devuelve {col_index: week_number}.
    Por defecto escanea todo el sheet para mayor robustez.
    """
    week_cols = {}
    patterns = [
        r"\bS\s*([1-9])\b",
        r"\bSemana\s*([1-9])\b",
        r"\bW\s*([1-9])\b",
    ]
    if scan_end_row is None:
        scan_end_row = ws.max_row
    for row in range(scan_start_row, scan_end_row + 1):
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=row, column=col).value
            if not isinstance(v, str):
                continue
            for pat in patterns:
                m = re.search(pat, v, flags=re.IGNORECASE)
                if m:
                    try:
                        wk = int(m.group(1))
                        if col not in week_cols:
                            week_cols[col] = wk
                    except Exception:
                        pass
                    break
    return week_cols

def create_perfect_template(num_days, original_path, output_path):
    """Crea una plantilla perfecta manteniendo exactamente la estructura original"""
    
    print(f"Creando plantilla perfecta para {num_days} días...")
    
    # Cargar plantilla original
    wb_original = openpyxl.load_workbook(original_path)
    ws_original = wb_original.active
    
    # Crear nuevo workbook
    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active
    
    # Copiar encabezado común (filas 1-9)
    copy_exact_structure(ws_original, ws_new, start_row=1, target_start_row=1, num_rows=9)
    
    # Copiar celdas combinadas del encabezado
    for merged_range in ws_original.merged_cells.ranges:
        if merged_range.max_row <= 9:
            ws_new.merge_cells(str(merged_range))
    
    current_row = 10
    
    # Para cada día, crear una sección única SIN duplicaciones
    # Solo copiamos la primera sección de día (filas 10-17) y omitimos las duplicadas (18-24, 25-31)
    for day_num in range(1, num_days + 1):
        # Copiar SOLO la primera sección de día (8 filas: 10-17)
        section_rows = 8  # Solo la primera sección sin duplicaciones
        copy_format_only(ws_original, ws_new, start_row=10, target_start_row=current_row, num_rows=section_rows)
        # Detectar columnas que representan semanas dentro de la sección
        # Detectar columnas de semana en todo el sheet para mayor certeza
        week_cols_map = detect_week_columns(ws_original, scan_start_row=1, scan_end_row=ws_original.max_row)
        
        # Procesar cada celda individualmente desde la fuente original (solo filas 10-17)
        for row_offset in range(section_rows):
            source_row = 10 + row_offset  # Solo filas 10-17
            target_row = current_row + row_offset
            
            for col in range(1, ws_original.max_column + 1):
                source_cell = ws_original.cell(row=source_row, column=col)
                target_cell = ws_new.cell(row=target_row, column=col)
                
                # Procesar SOLO si la celda original tiene contenido
                if source_cell.value and isinstance(source_cell.value, str):
                    original_value = source_cell.value
                    
                    # Aplicar reemplazos ÚNICOS y ESPECÍFICOS para este día
                    processed_value = original_value
                    
                    # Reemplazar EXACTAMENTE las variables originales
                    if "{{dia}}" in processed_value:
                        processed_value = processed_value.replace("{{dia}}", f"{{{{ dia_{day_num} }}}}")
                    
                    if "{{Ejercicio}}" in processed_value:
                        processed_value = processed_value.replace("{{Ejercicio}}", f"{{{{ ejercicio_dia_{day_num} }}}}")
                    
                    if "{{Series ejercicio}}" in processed_value:
                        # Si esta columna representa una semana, usar la variable _Sx
                        wk = week_cols_map.get(col)
                        if wk:
                            processed_value = processed_value.replace(
                                "{{Series ejercicio}}",
                                f"{{{{ series_dia_{day_num}_S{wk} }}}}"
                            )
                        else:
                            processed_value = processed_value.replace(
                                "{{Series ejercicio}}",
                                f"{{{{ series_dia_{day_num} }}}}"
                            )
                    
                    if "{{Repeticiones ejercicio}}" in processed_value:
                        wk = week_cols_map.get(col)
                        if wk:
                            processed_value = processed_value.replace(
                                "{{Repeticiones ejercicio}}",
                                f"{{{{ repeticiones_dia_{day_num}_S{wk} }}}}"
                            )
                        else:
                            processed_value = processed_value.replace(
                                "{{Repeticiones ejercicio}}",
                                f"{{{{ repeticiones_dia_{day_num} }}}}"
                            )
                    
                    # Asignar el valor procesado
                    target_cell.value = processed_value

                elif source_cell.value:
                    # Si no es string, copiar tal como está
                    target_cell.value = source_cell.value

        # Inyección explícita por fila:
        # Si una fila contiene el rótulo "Series ejercicio" en cualquier columna,
        # poblar todas las columnas de semana con {{ series_dia_{day_num}_Sx }}.
        # Similar para "Repeticiones ejercicio".
        # Buscamos en la fila de origen (10-17) y escribimos en la fila destino.
        for row_offset in range(section_rows):
            source_row = 10 + row_offset
            target_row = current_row + row_offset
            # Detectar tipo de fila
            row_text = " ".join(
                [str(ws_original.cell(row=source_row, column=c).value) for c in range(1, ws_original.max_column + 1)
                 if isinstance(ws_original.cell(row=source_row, column=c).value, str)]
            )
            is_series_row = "Series ejercicio" in row_text
            is_reps_row = "Repeticiones ejercicio" in row_text
            if is_series_row:
                for col, wk in week_cols_map.items():
                    ws_new.cell(row=target_row, column=col).value = f"{{{{ series_dia_{day_num}_S{wk} }}}}"
            if is_reps_row:
                for col, wk in week_cols_map.items():
                    ws_new.cell(row=target_row, column=col).value = f"{{{{ repeticiones_dia_{day_num}_S{wk} }}}}"

            # Asignación determinística basada en columnas que ya contienen variables del día
            # Mapea de izquierda a derecha: columnas con series -> S1,S2,... ; repeticiones -> S1,S2,...
            # Esto evita depender de encabezados de semana.
            # Detectar columnas con variables actuales
            series_cols = []
            reps_cols = []
            for c in range(1, ws_new.max_column + 1):
                val = ws_new.cell(row=target_row, column=c).value
                if isinstance(val, str):
                    if re.search(r"\{\{\s*series_dia_" + str(day_num) + r"\s*\}\}", val):
                        series_cols.append(c)
                    if re.search(r"\{\{\s*repeticiones_dia_" + str(day_num) + r"\s*\}\}", val):
                        reps_cols.append(c)

            # Reemplazar por versiones _Sx en orden
            for i, c in enumerate(sorted(series_cols)):
                ws_new.cell(row=target_row, column=c).value = f"{{{{ series_dia_{day_num}_S{i+1} }}}}"
            for i, c in enumerate(sorted(reps_cols)):
                ws_new.cell(row=target_row, column=c).value = f"{{{{ repeticiones_dia_{day_num}_S{i+1} }}}}"
        
        # Copiar celdas combinadas para este día (solo para filas 10-17)
        row_offset = current_row - 10
        for merged_range in ws_original.merged_cells.ranges:
            if 10 <= merged_range.min_row <= 17:  # Solo las de la primera sección (10-17)
                min_row = merged_range.min_row + row_offset
                max_row = merged_range.max_row + row_offset
                min_col = merged_range.min_col
                max_col = merged_range.max_col
                
                new_range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
                ws_new.merge_cells(new_range)
        
        current_row += section_rows
        
        # Agregar separación entre días (excepto el último)
        if day_num < num_days:
            current_row += 1
    
    # Guardar archivo
    wb_new.save(output_path)
    print(f"Plantilla perfecta guardada: {output_path}")
    print(f"Dimensiones finales: {ws_new.max_row} filas x {ws_new.max_column} columnas")
    
    wb_original.close()
    wb_new.close()

# Rutas
original_path = os.path.join('assets', 'templates', 'Plantilla Rutinas.xlsx')
base_path = os.path.abspath('.')

# Crear plantillas perfectas para 2, 3, 4 y 5 días
output_base = os.path.join('assets', 'templates')
os.makedirs(output_base, exist_ok=True)
templates_to_create = [
    (2, os.path.join(output_base, 'Plantilla_2_dias.xlsx')),
    (3, os.path.join(output_base, 'Plantilla_3_dias.xlsx')),
    (4, os.path.join(output_base, 'Plantilla_4_dias.xlsx')),
    (5, os.path.join(output_base, 'Plantilla_5_dias.xlsx'))
]

for num_days, output_path in templates_to_create:
    create_perfect_template(num_days, original_path, output_path)

print("\n=== TODAS LAS PLANTILLAS PERFECTAS CREADAS ===")
print("Estas plantillas mantienen exactamente:")
print("- Dimensiones de columnas y filas")
print("- Formato de fuentes, colores y alineación")
print("- Bordes y rellenos")
print("- Celdas combinadas")
print("- Variables Jinja2 adaptadas por día")