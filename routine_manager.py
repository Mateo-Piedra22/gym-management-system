#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
ROUTINE MANAGER - Sistema de Gestión de Plantillas de Rutinas
================================================================================

Este módulo maneja la carga de datos desde el programa hacia las plantillas Excel
usando xlsxtpl (Jinja2 + OpenPyXL) para un procesamiento más robusto y simple.

Características principales:
- Uso de xlsxtpl para plantillas con sintaxis Jinja2
- Carga secuencial de ejercicios por día
- Manejo correcto de series y repeticiones por semana
- Validación robusta de datos
- Logging detallado para debugging
- Soporte para múltiples plantillas (2, 3, 4, 5 días)

Autor: Sistema de Gestión de Gimnasio
Fecha: 2025
Versión: 60
"""

import os
import tempfile
import zipfile
import logging
from datetime import datetime, date
from typing import Dict, List, Any, Optional, Tuple
import json
from pathlib import Path
import re
import math
import subprocess

# Importaciones para Excel y PDF
from xlsxtpl.writerx import BookWriter
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.cell.cell import MergedCell
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl.drawing.image import Image as XLImage
try:
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
    from openpyxl.drawing.xdr import XDRPositiveSize2D
except Exception:
    AnchorMarker = OneCellAnchor = XDRPositiveSize2D = None
from openpyxl.utils import get_column_letter, column_index_from_string

# Importaciones del proyecto
from models import Rutina, RutinaEjercicio, Usuario, Ejercicio
from database import DatabaseManager
from utils import resource_path


class RoutineTemplateManager:
    """
    Gestor de plantillas de rutinas mejorado que usa xlsxtpl para un procesamiento
    más simple y robusto de las plantillas Excel.
    """
    
    def __init__(self, template_path: str = None, database_manager: DatabaseManager = None):
        """
        Inicializa el gestor de plantillas.
        
        Args:
            template_path: Ruta al archivo template Excel base (opcional)
            database_manager: Instancia del gestor de base de datos
        """
        # Configurar logging
        self.logger = logging.getLogger(__name__)
        
        # Rutas a las plantillas específicas por número de días (versiones perfectas)
        self.template_paths = {
            2: Path(resource_path(os.path.join("assets", "templates", "Plantilla_2_dias.xlsx"))),
            3: Path(resource_path(os.path.join("assets", "templates", "Plantilla_3_dias.xlsx"))), 
            4: Path(resource_path(os.path.join("assets", "templates", "Plantilla_4_dias.xlsx"))),
            5: Path(resource_path(os.path.join("assets", "templates", "Plantilla_5_dias.xlsx")))
        }
        
        # Directorios de salida con soporte de entorno y fallback a temporal si FS es de solo lectura
        pref_rutinas = os.environ.get("RUTINAS_DIR", "rutinas_exportadas")
        def _ensure_writable_dir(dir_path: str, fallback_subdir: str) -> Path:
            try:
                p = Path(dir_path)
                p.mkdir(exist_ok=True)
                return p
            except Exception:
                tmp_dir = Path(tempfile.gettempdir()) / fallback_subdir
                try:
                    tmp_dir.mkdir(exist_ok=True)
                except Exception:
                    tmp_dir = Path(tempfile.gettempdir())
                return tmp_dir

        self.output_dir_excel = _ensure_writable_dir(pref_rutinas, "rutinas_exportadas")
        self.output_dir_pdf = _ensure_writable_dir(pref_rutinas, "rutinas_exportadas")
        self.logger.info(f"Directorio Excel: '{self.output_dir_excel}'")
        self.logger.info(f"Directorio PDF: '{self.output_dir_pdf}'")
        
        # Guardar referencia al database manager
        self.db_manager = database_manager
        
        # Guardar template opcional (si lo pasan) aunque por defecto se seleccionan las plantillas perfectas
        self.custom_template_path = Path(template_path) if template_path else None
        
        self.logger.info("RoutineTemplateManager inicializado correctamente")
    
    def validate_routine_data(self, rutina: Rutina, usuario: Usuario, 
                            exercises_by_day: Dict[int, List[RutinaEjercicio]]) -> Tuple[bool, List[str]]:
        """
        Valida los datos de la rutina antes de procesarla.
        
        Args:
            rutina: Objeto Rutina
            usuario: Objeto Usuario
            exercises_by_day: Ejercicios organizados por día
            
        Returns:
            Tupla (es_válido, lista_errores)
        """
        errors = []
        
        try:
            # Validar rutina
            if not rutina:
                errors.append("Rutina no proporcionada")
            elif not getattr(rutina, 'nombre_rutina', None):
                errors.append("Rutina sin nombre")
            
            # Validar usuario
            if not usuario:
                errors.append("Usuario no proporcionado")
            elif not getattr(usuario, 'nombre', None):
                errors.append("Usuario sin nombre")
            
            # Validar ejercicios
            if not exercises_by_day:
                errors.append("No hay ejercicios en la rutina")
            else:
                total_exercises = sum(len(exercises) for exercises in exercises_by_day.values())
                if total_exercises == 0:
                    errors.append("La rutina no contiene ejercicios")
                
                # Validar cada día
                for day, exercises in exercises_by_day.items():
                    if not isinstance(day, int) or day < 1:
                        errors.append(f"Día inválido: {day}")
                    
                    for i, exercise in enumerate(exercises):
                        # Aceptar nombre proveniente de 'ejercicio.nombre' o del atributo personalizado 'nombre_ejercicio'
                        nombre_ok = False
                        try:
                            if getattr(exercise, 'nombre_ejercicio', None):
                                nombre_ok = True
                            elif getattr(exercise, 'ejercicio', None) and getattr(exercise.ejercicio, 'nombre', None):
                                nombre_ok = True
                        except Exception:
                            nombre_ok = False
                        if not nombre_ok:
                            errors.append(f"Ejercicio {i+1} del día {day} sin nombre")
            
            self.logger.info(f"Validación completada. Errores encontrados: {len(errors)}")
            return len(errors) == 0, errors
            
        except Exception as e:
            self.logger.error(f"Error durante validación: {e}")
            errors.append(f"Error de validación: {str(e)}")
            return False, errors

    def convert_excel_to_pdf(self, xlsx_path: str | Path, pdf_path: Optional[str | Path] = None) -> str:
        try:
            self.logger.info("Iniciando conversión de Excel a PDF")
            xlsx = Path(xlsx_path)
            if not xlsx.exists():
                raise FileNotFoundError(f"Archivo Excel no encontrado: {xlsx}")
            target_pdf = Path(pdf_path) if pdf_path else xlsx.with_suffix('.pdf')

            # En Vercel/serverless o en no-Windows, priorizamos el renderer Python (xlsx2pdf)
            prefer_python_renderer = (
                os.environ.get("VERCEL") == "1" or os.environ.get("PREFER_XLSX2PDF") == "1" or os.name != 'nt'
            )

            if prefer_python_renderer:
                # xlsx2pdf primero (API, luego CLI)
                try:
                    self.logger.info("Intentando conversión con xlsx2pdf (API)")
                    from xlsx2pdf import convert as xlsx2pdf_convert  # type: ignore
                    xlsx2pdf_convert(str(xlsx.resolve()), str(target_pdf.resolve()))
                    if target_pdf.exists():
                        self.logger.info(f"Conversión Excel→PDF exitosa con xlsx2pdf (API): {target_pdf}")
                        return str(target_pdf)
                    else:
                        raise RuntimeError("xlsx2pdf API no generó el PDF")
                except Exception as api_err:
                    self.logger.warning(f"xlsx2pdf API falló ({api_err}), probando CLI del módulo")
                    try:
                        import sys
                        py = sys.executable or "python"
                        cmd = [py, "-m", "xlsx2pdf", str(xlsx.resolve()), str(target_pdf.resolve())]
                        subprocess.run(cmd, check=True)
                        if target_pdf.exists():
                            self.logger.info(f"Conversión Excel→PDF exitosa con xlsx2pdf (CLI): {target_pdf}")
                            return str(target_pdf)
                        else:
                            raise RuntimeError("xlsx2pdf CLI no generó el PDF")
                    except Exception as cli_err:
                        self.logger.warning(f"xlsx2pdf también falló ({cli_err}), evaluando LibreOffice y fallback ReportLab")
                        # Si está disponible LibreOffice, intentar; de lo contrario, fallback
                        try:
                            import shutil
                            soffice = shutil.which("soffice") or shutil.which("soffice.exe")
                            if not soffice:
                                raise RuntimeError("LibreOffice no encontrado en el sistema")
                            outdir = str(target_pdf.parent.resolve())
                            cmd = [soffice, "--headless", "--convert-to", "pdf", "--outdir", outdir, str(xlsx.resolve())]
                            subprocess.run(cmd, check=True)
                            converted = xlsx.with_suffix('.pdf')
                            if converted.exists() and converted != target_pdf:
                                try:
                                    converted.replace(target_pdf)
                                except Exception:
                                    pass
                            self.logger.info(f"Conversión Excel→PDF exitosa con LibreOffice: {target_pdf}")
                            return str(target_pdf)
                        except Exception:
                            self.logger.info("Aplicando fallback con ReportLab para generar PDF desde Excel")
                            return self._excel_to_pdf_reportlab_fallback(xlsx, target_pdf)

            # Camino Windows tradicional: COM primero
            if os.name == 'nt':
                try:
                    import win32com.client as win32
                    excel = win32.DispatchEx("Excel.Application")
                    excel.Visible = False
                    excel.DisplayAlerts = False
                    wb = excel.Workbooks.Open(str(xlsx.resolve()))
                    try:
                        for ws in wb.Worksheets:
                            ps = ws.PageSetup
                            ps.Orientation = 2
                            ps.Zoom = False
                            ps.FitToPagesWide = 1
                            ps.FitToPagesTall = 0
                    except Exception:
                        pass
                    wb.ExportAsFixedFormat(0, str(target_pdf.resolve()))
                    wb.Close(False)
                    excel.Quit()
                    self.logger.info(f"Conversión Excel→PDF exitosa con COM: {target_pdf}")
                    return str(target_pdf)
                except Exception as e:
                    self.logger.warning(f"Fallo conversión con COM: {e}")

            # Como último recurso fuera del renderer Python, probar LibreOffice si existe
            try:
                import shutil
                soffice = shutil.which("soffice") or shutil.which("soffice.exe")
                if not soffice:
                    raise RuntimeError("LibreOffice no encontrado en el sistema")
                outdir = str(target_pdf.parent.resolve())
                cmd = [soffice, "--headless", "--convert-to", "pdf", "--outdir", outdir, str(xlsx.resolve())]
                subprocess.run(cmd, check=True)
                converted = xlsx.with_suffix('.pdf')
                if converted.exists() and converted != target_pdf:
                    try:
                        converted.replace(target_pdf)
                    except Exception:
                        pass
                self.logger.info(f"Conversión Excel→PDF exitosa con LibreOffice: {target_pdf}")
                return str(target_pdf)
            except Exception as lo_err:
                self.logger.info(f"LibreOffice no disponible: {lo_err}. Usando fallback ReportLab.")
                return self._excel_to_pdf_reportlab_fallback(xlsx, target_pdf)
        except Exception as e:
            self.logger.error(f"Conversión Excel→PDF fallida: {e}")
            raise e

    def _excel_to_pdf_reportlab_fallback(self, xlsx: Path, target_pdf: Path) -> str:
        """
        Fallback mejorado cuando no hay COM ni LibreOffice: renderiza la hoja
        de Excel a PDF respetando anchos de columna, alturas de fila, celdas
        combinadas, fondos, negritas y alineaciones básicas. Esto preserva la
        estética de las plantillas y evita el PDF "plano".
        """
        try:
            from openpyxl.utils import get_column_letter, column_index_from_string
            import re
            from typing import Optional
            wb = openpyxl.load_workbook(str(xlsx.resolve()), data_only=True)
            ws = wb.active

            # Determinar el rango útil (primer y último celda con contenido)
            min_row, min_col, max_row, max_col = None, None, 0, 0
            for r in ws.iter_rows():
                for c in r:
                    if (c.value is not None) or (c.fill and getattr(getattr(c.fill, 'fgColor', None), 'rgb', None)):
                        rr, cc = c.row, c.column
                        min_row = rr if (min_row is None or rr < min_row) else min_row
                        min_col = cc if (min_col is None or cc < min_col) else min_col
                        max_row = rr if rr > max_row else max_row
                        max_col = cc if cc > max_col else max_col
            if min_row is None:
                min_row, min_col, max_row, max_col = 1, 1, min(ws.max_row or 1, 120), min(ws.max_column or 1, 40)

            # Limitar por seguridad
            max_row = min(max_row, 200)
            max_col = min(max_col, 40)

            # Construir datos de tabla
            data = []
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                row_vals = []
                for cell in row:
                    val = cell.value
                    if val is None:
                        txt = ""
                    else:
                        try:
                            # Asegurar strings compatibles con ReportLab
                            txt = str(val)
                        except Exception:
                            txt = ""
                    row_vals.append(txt)
                data.append(row_vals)

            if not data:
                data = [["Documento vacío"], [f"Origen: {xlsx.name}"]]

            # Calcular anchos de columna aproximados según Excel
            raw_widths = []
            for c in range(min_col, max_col + 1):
                col_letter = get_column_letter(c)
                dim = ws.column_dimensions.get(col_letter)
                # Ancho Excel (caracteres) -> puntos aproximados
                excel_w = (getattr(dim, 'width', None) or 8.43)
                raw_widths.append(float(excel_w) * 7.2)  # ~7.2 pt por carácter

            # Calcular alturas de fila aproximadas
            raw_heights = []
            for r in range(min_row, max_row + 1):
                dim = ws.row_dimensions.get(r)
                excel_h = (getattr(dim, 'height', None) or 15.0)  # puntos
                raw_heights.append(float(excel_h))

            # Crear documento PDF
            doc = SimpleDocTemplate(
                str(target_pdf.resolve()),
                pagesize=landscape(A4),
                leftMargin=24,
                rightMargin=24,
                topMargin=28,
                bottomMargin=28,
            )
            page_width = landscape(A4)[0] - (doc.leftMargin + doc.rightMargin)
            sum_widths = sum(raw_widths) if raw_widths else page_width
            scale = page_width / sum_widths if sum_widths > 0 else 1.0
            col_widths = [w * scale for w in raw_widths]

            elements = []
            styles = getSampleStyleSheet()

            # Tabla principal con col/row sizes calculadas
            table = Table(data, colWidths=col_widths, rowHeights=raw_heights, repeatRows=1)
            ts = TableStyle([
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ])

            # Aplicar spans según celdas combinadas
            try:
                for rng in ws.merged_cells.ranges:
                    cr = rng.coord
                    # e.g., 'A1:C1'
                    m = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", cr)
                    if m:
                        c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
                        sc = column_index_from_string(c1) - min_col
                        sr = r1 - min_row
                        ec = column_index_from_string(c2) - min_col
                        er = r2 - min_row
                        ts.add("SPAN", (sc, sr), (ec, er))
            except Exception:
                pass

            # Función para obtener color hex de openpyxl
            def _hex_color(oc) -> Optional[str]:
                try:
                    if not oc:
                        return None
                    rgb = getattr(oc, 'rgb', None)
                    if rgb and len(rgb) == 8:  # ARGB
                        return f"#{rgb[2:]}"  # ignora alpha
                    rgb2 = getattr(oc, 'fgColor', None)
                    if rgb2 and getattr(rgb2, 'rgb', None):
                        v = rgb2.rgb
                        return f"#{v[2:]}" if len(v) == 8 else (f"#{v}" if len(v) == 6 else None)
                    return None
                except Exception:
                    return None

            # Aplicar estilos por celda: fondo, negrita, alineación
            for rr in range(min_row, max_row + 1):
                for cc in range(min_col, max_col + 1):
                    cell = ws.cell(row=rr, column=cc)
                    tr = rr - min_row
                    tc = cc - min_col
                    # Fondo
                    try:
                        fill = getattr(cell, 'fill', None)
                        if fill and getattr(fill, 'fill_type', None) and getattr(fill, 'fgColor', None):
                            hx = _hex_color(fill.fgColor)
                            if hx:
                                ts.add("BACKGROUND", (tc, tr), (tc, tr), colors.HexColor(hx))
                    except Exception:
                        pass
                    # Negrita
                    try:
                        font = getattr(cell, 'font', None)
                        if font and getattr(font, 'bold', False):
                            ts.add("FONTNAME", (tc, tr), (tc, tr), "Helvetica-Bold")
                        # Tamaño de fuente aproximado
                        fs = getattr(font, 'size', None)
                        if fs:
                            ts.add("FONTSIZE", (tc, tr), (tc, tr), float(fs))
                    except Exception:
                        pass
                    # Alineación
                    try:
                        align = getattr(cell, 'alignment', None)
                        h = getattr(align, 'horizontal', None)
                        if h:
                            ali = 'LEFT'
                            if str(h).lower() == 'center':
                                ali = 'CENTER'
                            elif str(h).lower() == 'right':
                                ali = 'RIGHT'
                            ts.add("ALIGN", (tc, tr), (tc, tr), ali)
                    except Exception:
                        pass

            table.setStyle(ts)
            elements.append(table)

            doc.build(elements)
            self.logger.info(f"Conversión Excel→PDF (renderer Python) generada: {target_pdf}")
            return str(target_pdf)
        except Exception as e:
            self.logger.error(f"Error en fallback ReportLab, sin generar PDF genérico: {e}")
            # Forzar error para garantizar uso obligatorio de plantillas y flujo Excel→PDF
            raise e
    
    def _select_template_by_days(self, num_days: int) -> Path:
        """
        Selecciona la plantilla correcta según el número de días.
        
        Args:
            num_days: Número de días de la rutina
            
        Returns:
            Path a la plantilla seleccionada
        """
        # Mapear número de días a plantilla disponible
        if num_days <= 2:
            template_key = 2
        elif num_days <= 3:
            template_key = 3
        elif num_days <= 4:
            template_key = 4
        else:
            template_key = 5
        
        template_path = self.template_paths[template_key]
        
        if not template_path.exists():
            self.logger.warning(f"Plantilla {template_path} no encontrada, usando plantilla de 3 días")
            template_path = self.template_paths[3]
        
        self.logger.info(f"Plantilla seleccionada: {template_path} para {num_days} días")
        return template_path
    
    def _prepare_template_data(self, rutina: Rutina, usuario: Usuario, 
                             exercises_by_day: Dict[int, List[RutinaEjercicio]], 
                             weeks: int = 1) -> Dict[str, Any]:
        """
        Prepara los datos para la plantilla Excel con variables específicas para plantillas perfectas.
        
        Args:
            rutina: Objeto Rutina
            usuario: Objeto Usuario  
            exercises_by_day: Diccionario con ejercicios por día
            weeks: Número de semanas
            
        Returns:
            Diccionario con datos formateados para la plantilla
        """
        self.logger.info("Preparando datos para plantilla Excel con variables específicas por día")
        
        # Determinar semana actual de forma robusta
        current_week = getattr(rutina, 'semana', None) or getattr(rutina, 'semana_actual', None) or 1
        try:
            current_week = int(current_week)
        except Exception:
            current_week = 1
        if weeks is None or weeks <= 0:
            weeks = 1
        if current_week < 1:
            current_week = 1
        if current_week > weeks:
            current_week = weeks
        
        # Datos básicos del usuario y rutina
        template_data = {
            'nombre_usuario': getattr(usuario, 'nombre', '') or '',
            'nombre_completo': getattr(usuario, 'nombre', '') or '',
            'Nombre': getattr(usuario, 'nombre', '') or '',
            'rutina_nombre': getattr(rutina, 'nombre_rutina', '') or '',
            'nombre_rutina': getattr(rutina, 'nombre_rutina', '') or '',
            'Rutina': getattr(rutina, 'nombre_rutina', '') or '',
            'fecha_creacion': datetime.now().strftime("%d/%m/%Y"),
            'fecha': datetime.now().strftime("%d/%m"),
            'año': datetime.now().year,
            'semana': f"Semana {current_week}",  # Ajuste: usar current_week
            'semana_num': current_week,          # Valor numérico disponible si lo requieren plantillas
            'logo_gimnasio': '__GYM_LOGO__',
            # Datos de contacto
            'dni': getattr(usuario, 'dni', '') or '',
            'DNI': getattr(usuario, 'dni', '') or '',
            'telefono': getattr(usuario, 'telefono', '') or '',
            'Teléfono': getattr(usuario, 'telefono', '') or '',
        }

        # Estructuras anidadas para compatibilidad con plantillas que usan .get
        template_data['usuario'] = {
            'nombre': getattr(usuario, 'nombre', '') or '',
            'dni': getattr(usuario, 'dni', '') or '',
            'telefono': getattr(usuario, 'telefono', '') or '',
        }
        template_data['rutina'] = {
            'nombre': getattr(rutina, 'nombre_rutina', '') or '',
            'fecha': template_data['fecha'],
            'semana': template_data['semana_num'],
            'semana_texto': template_data['semana'],  # "Semana x"
            'Semana': template_data['semana'],        # Alias "Semana x"
        }
        
        # Variables específicas para plantillas perfectas (formato requerido)
        template_data.update({
            'Nombre completo': getattr(usuario, 'nombre', '') or '',
            'Nombre_completo': getattr(usuario, 'nombre', '') or '',
            'Año': datetime.now().year,
            'Fecha': datetime.now().strftime("%d/%m"),
            'Logo Gimnasio': '__GYM_LOGO__',
            'Logo_Gimnasio': '__GYM_LOGO__',
            'Semana': f"Semana {current_week}",  # Ajuste: usar current_week
            'current_week': current_week,        # Valor numérico de semana actual
            'Week': current_week,
            'Semana_num': current_week,
            'Semana_numero': current_week,
        })
        
        # Generar variables específicas por día para plantillas perfectas
        for day_num in sorted(exercises_by_day.keys()):
            exercises = exercises_by_day[day_num]
            
            # Variables para el día actual
            dia_key = f"dia_{day_num}"
            ejercicio_key = f"ejercicio_dia_{day_num}"
            series_key = f"series_dia_{day_num}"
            repeticiones_key = f"repeticiones_dia_{day_num}"
            
            # Nombre del día
            template_data[dia_key] = f"Dia {day_num}"  # Ajuste: texto "Dia x" sin tilde
            template_data[f" {dia_key}"] = f"Dia {day_num}"  # Con espacio para compatibilidad
            
            # Listas para ejercicios, series y repeticiones del día
            ejercicios_dia = []
            series_dia = []
            repeticiones_dia = []
            
            # Listas para cada semana específica
            series_por_semana = {f"S{i}": [] for i in range(1, weeks + 1)}
            repeticiones_por_semana = {f"S{i}": [] for i in range(1, weeks + 1)}
            
            for ejercicio in exercises:
                # Nombre del ejercicio preferentemente desde atributo personalizado si existe
                nombre_ejercicio = (
                    getattr(ejercicio, 'nombre_ejercicio', None)
                    or (ejercicio.ejercicio.nombre if getattr(ejercicio, 'ejercicio', None) else None)
                    or f"Ejercicio {getattr(ejercicio, 'ejercicio_id', '')}"
                )
                ejercicios_dia.append(nombre_ejercicio)
                
                # Parsear series y repeticiones por semana
                series_semanas = self._parse_weekly_values(getattr(ejercicio, 'series', ''), weeks)
                repeticiones_semanas = self._parse_weekly_values(getattr(ejercicio, 'repeticiones', ''), weeks)
                
                # Usar exclusivamente la semana actual para las variables básicas; si falta, dejar vacío
                idx_sem_actual = max(0, min(weeks - 1, current_week - 1))
                series_valor = series_semanas[idx_sem_actual] if 0 <= idx_sem_actual < len(series_semanas) else ""
                repeticiones_valor = repeticiones_semanas[idx_sem_actual] if 0 <= idx_sem_actual < len(repeticiones_semanas) else ""
                
                series_dia.append(series_valor)
                repeticiones_dia.append(repeticiones_valor)
                
                # Agregar valores por semana específica SIN rellenar vacíos desde valores planos u otras semanas
                for week_idx in range(weeks):
                    semana_key = f"S{week_idx + 1}"
                    series_semana = series_semanas[week_idx] if week_idx < len(series_semanas) else ""
                    reps_semana = repeticiones_semanas[week_idx] if week_idx < len(repeticiones_semanas) else ""
                    
                    series_por_semana[semana_key].append(series_semana)
                    repeticiones_por_semana[semana_key].append(reps_semana)
            
            # Guardar listas crudas para distribución por filas (fallback)
            template_data[f"lista_ejercicios_dia_{day_num}"] = ejercicios_dia
            template_data[f"lista_series_dia_{day_num}"] = series_dia
            template_data[f"lista_repeticiones_dia_{day_num}"] = repeticiones_dia
            for semana_key, lista in series_por_semana.items():
                template_data[f"lista_series_dia_{day_num}_{semana_key}"] = lista
            for semana_key, lista in repeticiones_por_semana.items():
                template_data[f"lista_repeticiones_dia_{day_num}_{semana_key}"] = lista
            
            # Convertir listas a strings para las plantillas
            template_data[ejercicio_key] = '\n'.join(map(str, ejercicios_dia))
            template_data[series_key] = '\n'.join(map(str, series_dia))
            template_data[repeticiones_key] = '\n'.join(map(str, repeticiones_dia))
            
            # Variables con espacio para compatibilidad con plantillas
            template_data[f" {ejercicio_key}"] = '\n'.join(map(str, ejercicios_dia))
            template_data[f" {series_key}"] = '\n'.join(map(str, series_dia))
            template_data[f" {repeticiones_key}"] = '\n'.join(map(str, repeticiones_dia))
            
            # Generar variables específicas por semana para cada día
            for semana_key in series_por_semana.keys():
                # Variables de series por semana
                series_semana_key = f"series_dia_{day_num}_{semana_key}"
                template_data[series_semana_key] = '\n'.join(map(str, series_por_semana[semana_key]))
                template_data[f" {series_semana_key}"] = '\n'.join(map(str, series_por_semana[semana_key]))
                
                # Variables de repeticiones por semana
                reps_semana_key = f"repeticiones_dia_{day_num}_{semana_key}"
                template_data[reps_semana_key] = '\n'.join(map(str, repeticiones_por_semana[semana_key]))
                template_data[f" {reps_semana_key}"] = '\n'.join(map(str, repeticiones_por_semana[semana_key]))
            
            self.logger.info(f"Generadas variables para {dia_key}: {len(ejercicios_dia)} ejercicios, {weeks} semanas")
        
        # Mantener estructura de días para compatibilidad con código existente
        template_data['dias'] = []
        for day_num in sorted(exercises_by_day.keys()):
            exercises = exercises_by_day[day_num]
            day_data = {
                'numero': day_num,
                'nombre': f"Día {day_num}",
                'ejercicios': []
            }
            
            for ejercicio in exercises:
                series_semanas = self._parse_weekly_values(getattr(ejercicio, 'series', ''), weeks)
                repeticiones_semanas = self._parse_weekly_values(getattr(ejercicio, 'repeticiones', ''), weeks)
                
                exercise_data = {
                    'nombre': (
                        getattr(ejercicio, 'nombre_ejercicio', None)
                        or (ejercicio.ejercicio.nombre if getattr(ejercicio, 'ejercicio', None) else f"Ejercicio {getattr(ejercicio, 'ejercicio_id', '')}")
                    ),
                    'series': getattr(ejercicio, 'series', ''),
                    'repeticiones': getattr(ejercicio, 'repeticiones', ''),
                    'series_semanas': series_semanas,
                    'repeticiones_semanas': repeticiones_semanas,
                    'notas': getattr(ejercicio, 'notas', '') or ""
                }
                day_data['ejercicios'].append(exercise_data)
            
            template_data['dias'].append(day_data)
        
        self.logger.info(f"Datos preparados para {len(template_data['dias'])} días con variables específicas por día")
        return template_data
    
    def _parse_weekly_values(self, value_string: str, weeks: int) -> List[str]:
        """
        Parsea una cadena de valores separados por comas para múltiples semanas.
        Args:
            value_string: Cadena con valores separados por comas
            weeks: Número de semanas esperadas
        Returns:
            Lista de valores por semana. Si faltan valores, se completan con cadenas vacías.
        """
        # Manejar None o valores no cadena de forma robusta
        if value_string is None:
            return [""] * weeks
        try:
            raw_str = str(value_string)
        except Exception:
            raw_str = ""
        if raw_str.strip() == "":
            return [""] * weeks
        # Dividir por comas y limpiar espacios
        values = [v.strip() for v in raw_str.split(",")]
        # Si hay menos valores que semanas, rellenar con cadenas vacías ("")
        # No repetir el último valor.
        while len(values) < weeks:
            values.append("")
        # Si hay más valores que semanas, tomar solo los primeros
        return values[:weeks]
    
    def _sanitize_template_for_jinja(self, template_path: Path) -> Path:
        """
        Crea una copia temporal de la plantilla reemplazando variables con espacios
        por equivalentes válidos en Jinja (espacios -> guiones bajos), p.ej.:
        {{Nombre completo}} -> {{ Nombre_completo }}
        {{Logo Gimnasio}}  -> {{ Logo_Gimnasio }}
        No toca expresiones complejas (con [], (), '.', ':', comillas, etc.).
        """
        try:
            wb = openpyxl.load_workbook(str(template_path), data_only=False)
            for sh in wb.worksheets:
                for row in sh.iter_rows():
                    for cell in row:
                        v = cell.value
                        if isinstance(v, str) and '{{' in v and '}}' in v:
                            def _repl(m):
                                inner = m.group(1)
                                if any(ch in inner for ch in "[]()'\".:}|{"):
                                    return "{{ " + inner.strip() + " }}"
                                normalized = re.sub(r"\s+", "_", inner.strip())
                                return "{{ " + normalized + " }}"
                            new_v = re.sub(r"\{\{\s*([^}]+?)\s*\}\}", _repl, v)
                            if new_v != v:
                                cell.value = new_v
            sanitized_path = self.output_dir_excel / f"__sanitized_{template_path.name}"
            wb.save(str(sanitized_path))
            wb.close()
            self.logger.info(f"Plantilla sanitizada para Jinja: {sanitized_path}")
            return sanitized_path
        except Exception as e:
            self.logger.warning(f"No se pudo sanitizar plantilla para Jinja ({e}), usando original")
            return template_path

    def generate_routine_excel(self, rutina: Rutina, usuario: Usuario,
                             exercises_by_day: Dict[int, List[RutinaEjercicio]], 
                             output_path: str = None, weeks: int = 1) -> str:
        """
        Genera un archivo Excel de rutina usando xlsxtpl.
        
        Args:
            rutina: Objeto Rutina
            usuario: Objeto Usuario
            exercises_by_day: Diccionario con ejercicios por día
            output_path: Ruta de salida personalizada
            weeks: Número de semanas
            
        Returns:
            Ruta del archivo generado
        """
        try:
            self.logger.info("Iniciando generación de Excel de rutina")
            
            # Validar datos
            is_valid, errors = self.validate_routine_data(rutina, usuario, exercises_by_day)
            if not is_valid:
                error_msg = f"Datos de rutina inválidos: {', '.join(errors)}"
                self.logger.error(error_msg)
                raise ValueError(error_msg)
            
            # Clamp days to max 5 and exercises per day to max 8, luego seleccionar plantilla
            try:
                if exercises_by_day:
                    trimmed = {}
                    for d in sorted(exercises_by_day.keys())[:5]:
                        day_list = exercises_by_day.get(d) or []
                        trimmed[d] = day_list[:8]
                    exercises_by_day = trimmed
            except Exception:
                pass
            num_days = len(exercises_by_day)
            template_path = self._select_template_by_days(num_days)
            
            if not template_path.exists():
                raise FileNotFoundError(f"Plantilla no encontrada: {template_path}")
            
            # Generar nombre de archivo si no se proporciona
            if not output_path:
                fecha_str = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"rutina_{usuario.nombre.replace(' ', '_')}_{num_days}dias_{fecha_str}.xlsx"
                output_path = self.output_dir_excel / filename
            else:
                output_path = Path(output_path)
            
            # Preparar datos para la plantilla
            # Inferir semanas efectivas a partir de los datos si el parámetro es impreciso
            weeks_effective = weeks if isinstance(weeks, int) and weeks > 0 else 0
            try:
                max_weeks_detected = 0
                for day_exs in (exercises_by_day or {}).values():
                    for ej in (day_exs or []):
                        for raw in [getattr(ej, 'series', ''), getattr(ej, 'repeticiones', '')]:
                            s = str(raw) if raw is not None else ''
                            if s.strip():
                                count = len([v.strip() for v in s.split(',')])
                                if count > max_weeks_detected:
                                    max_weeks_detected = count
                if weeks_effective <= 0:
                    weeks_effective = max_weeks_detected or 1
                else:
                    weeks_effective = max(weeks_effective, max_weeks_detected or 1)
            except Exception:
                weeks_effective = weeks if isinstance(weeks, int) and weeks > 0 else 1
            # Clamp weeks to maximum of 4
            try:
                weeks_effective = max(1, min(int(weeks_effective), 4))
            except Exception:
                weeks_effective = 1
            
            template_data = self._prepare_template_data(rutina, usuario, exercises_by_day, weeks_effective)
            
            self.logger.info(f"Procesando plantilla: {template_path}")
            self.logger.info(f"Datos preparados con {len(template_data['dias'])} días")
            
            # Si la plantilla repite placeholders por fila, usar fallback para distribuir por filas
            if self._template_has_repeated_day_placeholders(template_path):
                self.logger.info("Detectadas variables por día repetidas en plantilla; usando método fallback de distribución por filas")
                return self._generate_excel_fallback(template_path, template_data, Path(output_path))
        
            # Usar xlsxtpl para procesar la plantilla
            try:
                # Sanitizar variables inválidas para Jinja y crear el writer
                sanitized_template = self._sanitize_template_for_jinja(template_path)
                writer = BookWriter(str(sanitized_template))
                # Endurecer el entorno Jinja: evitar variables indefinidas silenciosas y globals peligrosos
                try:
                    from jinja2 import StrictUndefined as _StrictUndefined
                    writer.jinja_env.undefined = _StrictUndefined
                    # Asegurar que no haya helpers peligrosos expuestos
                    for _k in list(getattr(writer.jinja_env, 'globals', {}).keys()):
                        if _k in ('dir', 'getattr', '__import__', 'eval', 'exec', 'open'):
                            try:
                                del writer.jinja_env.globals[_k]
                            except Exception:
                                pass
                except Exception:
                    pass
                
                # Preparar datos para xlsxtpl (incluir estructura completa por si hay bucles Jinja)
                xlsxtpl_data = template_data
                
                # Renderizar la plantilla con los datos
                writer.render_book(payloads=[xlsxtpl_data])
                
                # Guardar el archivo
                writer.save(str(output_path))
                
                # Validar que el archivo generado sea un XLSX (ZIP) válido
                try:
                    if not zipfile.is_zipfile(str(output_path)):
                        raise RuntimeError("Archivo Excel generado inválido (no es ZIP).")
                except Exception as _zip_err:
                    # Forzar manejo en el catch general para activar fallback
                    raise _zip_err
                
                # Aplicar anchos de columna en todas las hojas (semanas) tras guardar
                try:
                    wb = openpyxl.load_workbook(str(output_path))
                    # Renombrar encabezados 'Series' -> 'Ser' y 'Repeticiones' -> 'Rep' en todas las hojas
                    try:
                        for ws in wb.worksheets:
                            for row in ws.iter_rows():
                                for cell in row:
                                    if isinstance(cell, MergedCell):
                                        continue
                                    v = cell.value
                                    if isinstance(v, str):
                                        t = v.strip()
                                        if t == "Series":
                                             cell.value = "Ser"
                                        elif t == "Repeticiones":
                                             cell.value = "Rep"
                    except Exception:
                        pass
                    # Insertar imágenes de logo donde corresponda
                    try:
                        logo_path = resource_path(os.path.join('assets', 'gym_logo.png'))
                        if os.path.exists(logo_path):
                            SENTINEL = '__GYM_LOGO__'
                            for ws in wb.worksheets:
                                for row in ws.iter_rows():
                                    for cell in row:
                                        if isinstance(cell, MergedCell):
                                            continue
                                        v = cell.value
                                        if isinstance(v, str) and (SENTINEL in v or ('{{' in v and '}}' in v and (('logo' in v.lower()) and ('gimnasio' in v.lower())))):
                                            try:
                                                cell.value = ''
                                            except Exception:
                                                pass
                                            try:
                                                self._insert_logo_in_cell(ws, cell, logo_path)
                                            except Exception:
                                                pass
                    except Exception:
                        pass
                    _apply_column_widths_all_sheets(self, wb, 'A', 'K', 4.57, pixel_width=37)
                    _apply_column_widths_all_sheets(self, wb, 'L', 'AA', 3.71, pixel_width=31)
                    # Ocultar filas en blanco específicas para preservar bordes y fusiones
                    try:
                        for ws in wb.worksheets:
                            for r in (18, 27, 36, 45):
                                try:
                                    ws.row_dimensions[r].hidden = True
                                except Exception:
                                    pass
                    except Exception:
                        pass
                    # Forzar altura de filas a 17px (≈12.75pt) en todas las hojas
                    try:
                        for ws in wb.worksheets:
                            try:
                                if getattr(ws, 'sheet_format', None) is not None:
                                    ws.sheet_format.defaultRowHeight = 12.75
                            except Exception:
                                pass
                            try:
                                max_r = getattr(ws, 'max_row', 0) or 0
                                for r in range(1, max_r + 1):
                                    ws.row_dimensions[r].height = 12.75
                            except Exception:
                                pass
                    except Exception:
                        pass
                    wb.save(str(output_path))
                    wb.close()
                except Exception:
                    pass

                # Reabrir y reforzar anchos de columna para evitar normalizaciones de Excel (xlsxtpl)
                try:
                    wb3 = openpyxl.load_workbook(str(output_path))
                    try:
                        _apply_column_widths_all_sheets(self, wb3, 'A', 'K', 4.57, pixel_width=37)
                        _apply_column_widths_all_sheets(self, wb3, 'L', 'AA', 3.71, pixel_width=31)
                        # Reforzar altura de filas 17px (≈12.75pt)
                        try:
                            for ws in wb3.worksheets:
                                try:
                                    if getattr(ws, 'sheet_format', None) is not None:
                                        ws.sheet_format.defaultRowHeight = 12.75
                                except Exception:
                                    pass
                                try:
                                    max_r = getattr(ws, 'max_row', 0) or 0
                                    for r in range(1, max_r + 1):
                                        ws.row_dimensions[r].height = 12.75
                                except Exception:
                                    pass
                        except Exception:
                            pass
                    except Exception:
                        pass
                    wb3.save(str(output_path))
                    try:
                        wb3.close()
                    except Exception:
                        pass
                except Exception:
                    pass
                
                # Intentar eliminar plantilla temporal
                try:
                    if sanitized_template != template_path and os.path.exists(sanitized_template):
                        os.remove(sanitized_template)
                except Exception:
                    pass
                
                self.logger.info(f"Rutina Excel generada exitosamente: {output_path}")
                return str(output_path)
                 
            except Exception as template_error:
                self.logger.exception(f"Error procesando plantilla con xlsxtpl: {template_error}")
                
                # Fallback: usar método tradicional con openpyxl
                self.logger.info("Usando método fallback con openpyxl")
                return self._generate_excel_fallback(template_path, template_data, output_path)
            
        except Exception as e:
            self.logger.error(f"Error al generar Excel de rutina: {e}")
            raise e
    
    def _template_has_repeated_day_placeholders(self, template_path: Path) -> bool:
        """
        Detecta si la plantilla contiene placeholders de día (ejercicio/series/repeticiones)
        repetidos en múltiples filas dentro de la MISMA hoja. Esto evita falsos positivos
        cuando un placeholder aparece varias veces en la misma fila o en distintas hojas.
        """
        try:
            wb = openpyxl.load_workbook(str(template_path), data_only=False)
            patterns = {
                'ejercicio': re.compile(r"\{\{\s*ejercicio_dia_(\d+)\s*\}\}"),
                'series': re.compile(r"\{\{\s*series_dia_(\d+)(?:_S\d+)?\s*\}\}"),
                'repeticiones': re.compile(r"\{\{\s*repeticiones_dia_(\d+)(?:_S\d+)?\s*\}\}")
            }
            for sh in wb.worksheets:
                # Mapear (tipo, dia) -> conjunto de filas distintas donde aparece en ESTA hoja
                rows_by_key = {}
                for row in sh.iter_rows():
                    for cell in row:
                        v = cell.value
                        if isinstance(v, str) and '{{' in v and '}}' in v:
                            for kind, rx in patterns.items():
                                for match in rx.findall(v):
                                    key = (kind, int(match))
                                    rows_by_key.setdefault(key, set()).add(cell.row)
                # Si en esta hoja un placeholder aparece en más de una fila, requiere fallback
                for rows in rows_by_key.values():
                    if len(rows) > 1:
                        wb.close()
                        return True
            wb.close()
            return False
        except Exception as e:
            self.logger.warning(f"No se pudo analizar plantilla para repetición de placeholders: {e}")
            return False

    def _generate_excel_fallback(self, template_path: Path, template_data: Dict[str, Any], 
                               output_path: Path) -> str:
        """
        Método fallback para generar Excel usando openpyxl directamente.
        
        Args:
            template_path: Ruta a la plantilla
            template_data: Datos preparados
            output_path: Ruta de salida
            
        Returns:
            Ruta del archivo generado
        """
        try:
            # Cargar plantilla
            workbook = openpyxl.load_workbook(str(template_path))
            sheet = workbook.active
            
            # Reemplazar variables básicas
            self._replace_basic_variables(sheet, template_data)
            
            # Cargar ejercicios de forma secuencial
            self._load_exercises_sequential(sheet, template_data['dias'], template_data.get('current_week'))
            
            # Aplicar anchos de columna A..K = 4.57 y L..AA = 3.71 en todas las hojas antes de guardar
            try:
                _apply_column_widths_all_sheets(self, workbook, 'A', 'K', 4.57, pixel_width=37)
                _apply_column_widths_all_sheets(self, workbook, 'L', 'AA', 3.71, pixel_width=31)
            except Exception:
                pass
            # Insertar imágenes de logo donde corresponda
            try:
                logo_path = resource_path(os.path.join('assets', 'gym_logo.png'))
                if os.path.exists(logo_path):
                    SENTINEL = '__GYM_LOGO__'
                    for ws in workbook.worksheets:
                        for row in ws.iter_rows():
                            for cell in row:
                                if isinstance(cell, MergedCell):
                                    continue
                                v = cell.value
                                if isinstance(v, str) and (SENTINEL in v or ('{{' in v and '}}' in v and (('logo' in v.lower()) and ('gimnasio' in v.lower())))):
                                    try:
                                        cell.value = ''
                                    except Exception:
                                        pass
                                    try:
                                        self._insert_logo_in_cell(ws, cell, logo_path)
                                    except Exception:
                                        pass
            except Exception:
                pass
            # Renombrar encabezados 'Series' -> 'Ser' y 'Repeticiones' -> 'Rep' en todas las hojas antes de guardar
            try:
                for ws in workbook.worksheets:
                    for row in ws.iter_rows():
                        for cell in row:
                            if isinstance(cell, MergedCell):
                                continue
                            v = cell.value
                            if isinstance(v, str):
                                t = v.strip()
                                if t == "Series":
                                    cell.value = "Ser"
                                elif t == "Repeticiones":
                                    cell.value = "Rep"
            except Exception:
                pass
            # Forzar altura de filas a 17px (≈12.75pt) en todas las hojas antes de guardar
            try:
                for ws in workbook.worksheets:
                    try:
                        if getattr(ws, 'sheet_format', None) is not None:
                            ws.sheet_format.defaultRowHeight = 12.75
                    except Exception:
                        pass
                    try:
                        max_r = getattr(ws, 'max_row', 0) or 0
                        for r in range(1, max_r + 1):
                            ws.row_dimensions[r].height = 12.75
                    except Exception:
                        pass
            except Exception:
                pass
            # Ocultar filas en blanco específicas para preservar bordes y fusiones
            try:
                for ws in workbook.worksheets:
                    for r in (18, 27, 36, 45):
                        try:
                            ws.row_dimensions[r].hidden = True
                        except Exception:
                            pass
            except Exception:
                pass
            # Guardar archivo
            workbook.save(str(output_path))
            # Cerrar workbook para liberar el archivo en Windows
            try:
                workbook.close()
            except Exception:
                pass

            # Reabrir y reforzar anchos de columna para evitar normalizaciones de Excel
            try:
                wb2 = openpyxl.load_workbook(str(output_path))
                try:
                    _apply_column_widths_all_sheets(self, wb2, 'A', 'K', 4.57, pixel_width=37)
                    _apply_column_widths_all_sheets(self, wb2, 'L', 'AA', 3.71, pixel_width=31)
                except Exception:
                    pass
                # Reforzar altura de filas a 17px (≈12.75pt) tras reabrir
                try:
                    for ws in wb2.worksheets:
                        try:
                            if getattr(ws, 'sheet_format', None) is not None:
                                ws.sheet_format.defaultRowHeight = 12.75
                        except Exception:
                            pass
                        try:
                            max_r = getattr(ws, 'max_row', 0) or 0
                            for r in range(1, max_r + 1):
                                ws.row_dimensions[r].height = 12.75
                        except Exception:
                            pass
                except Exception:
                    pass
                wb2.save(str(output_path))
                try:
                    wb2.close()
                except Exception:
                    pass
            except Exception:
                pass
            
            self.logger.info(f"Excel generado con método fallback: {output_path}")
            return str(output_path)
            
        except Exception as e:
            self.logger.error(f"Error en método fallback: {e}")
            raise e
    
    def _replace_basic_variables(self, sheet, template_data: Dict[str, Any]):
        """
        Reemplaza variables básicas y específicas por día en la hoja de Excel.
        
        Args:
            sheet: Hoja de Excel
            template_data: Datos de la plantilla
        """
        # Mapeo de variables comunes
        variable_mapping = {
            '{{nombre_completo}}': template_data.get('nombre_completo', ''),
            '{{Nombre completo}}': template_data.get('Nombre completo', ''),
            '{{Nombre}}': template_data.get('Nombre', ''),
            '{{rutina_nombre}}': template_data.get('rutina_nombre', ''),
            '{{nombre_rutina}}': template_data.get('nombre_rutina', ''),
            '{{Rutina}}': template_data.get('Rutina', ''),
            '{{fecha}}': template_data.get('fecha', ''),
            '{{Fecha}}': template_data.get('Fecha', ''),
            '{{año}}': template_data.get('año', ''),
            '{{Año}}': template_data.get('Año', ''),
            # '{{semana}}' y '{{Semana}}' se manejan de forma contextual por columna; evitamos reemplazo global aquí
            # '{{semana}}': template_data.get('semana', ''),
            # '{{Semana}}': template_data.get('Semana', ''),
            '{{semana_num}}': template_data.get('semana_num', ''),
            '{{Semana_num}}': template_data.get('Semana_num', ''),
            '{{Semana_numero}}': template_data.get('Semana_numero', ''),
            '{{current_week}}': template_data.get('current_week', ''),
            '{{Week}}': template_data.get('Week', ''),
            '{{logo_gimnasio}}': template_data.get('logo_gimnasio', ''),
            '{{Logo Gimnasio}}': template_data.get('Logo Gimnasio', ''),
            '{{Logo_Gimnasio}}': template_data.get('Logo_Gimnasio', ''),
            '{{dni}}': template_data.get('dni', ''),
            '{{DNI}}': template_data.get('DNI', ''),
            '{{telefono}}': template_data.get('telefono', ''),
            '{{Teléfono}}': template_data.get('Teléfono', ''),
        }
        
        # También contemplar variantes con espacios (inmediatamente después de '{{' y antes de '}}')
        base_keys_with_space = {}
        base_keys_trailing_space = {}
        for k, v in list(variable_mapping.items()):
            # Insertar variante con espacio inmediatamente después de '{{'
            if k.startswith('{{') and not k.startswith('{{ '):
                base_keys_with_space['{{ ' + k[2:]] = v
            # Insertar variante con espacio antes de '}}'
            if k.endswith('}}') and not k.endswith(' }}'):
                base_keys_trailing_space[k[:-2] + ' }}'] = v
        variable_mapping.update(base_keys_with_space)
        variable_mapping.update(base_keys_trailing_space)
        
        # Agregar variables específicas por día (sustitución global de texto)
        for key, value in template_data.items():
            if key.startswith(('dia_', 'ejercicio_dia_', 'series_dia_', 'repeticiones_dia_')):
                # Sin espacios
                variable_mapping[f'{{{{{key}}}}}'] = str(value)
                # Espacio después de '{{'
                variable_mapping[f'{{{{ {key}}}}}'] = str(value)
                # Espacio antes de '}}'
                variable_mapping[f'{{{{{key} }}}}'] = str(value)
                # Espacio en ambos lados
                variable_mapping[f'{{{{ {key} }}}}'] = str(value)
        
        # Preparar estructuras para distribución por filas (cuando la celda es exactamente el placeholder)
        dias_list = template_data.get('dias', []) or []
        ejercicios_por_dia: Dict[int, List[Dict[str, Any]]] = {}
        puntero_por_dia: Dict[int, int] = {}
        for d in dias_list:
            # Aceptar tanto 'numero' (actual) como 'dia' (anterior)
            day_num = d.get('numero') if isinstance(d, dict) else None
            if day_num is None:
                day_num = d.get('dia') if isinstance(d, dict) else None
            if isinstance(day_num, int):
                ejercicios_por_dia[day_num] = d.get('ejercicios', []) or []
                puntero_por_dia[day_num] = 0
        
        # Calcular número de semanas disponible para acotar current_week_idx
        try:
            weeks_count = 1
            for d in dias_list:
                for ej in d.get('ejercicios', []) or []:
                    sw = ej.get('series_semanas') or []
                    rw = ej.get('repeticiones_semanas') or []
                    weeks_count = max(weeks_count, len(sw) if isinstance(sw, list) else 0, len(rw) if isinstance(rw, list) else 0)
            if weeks_count <= 0:
                weeks_count = 1
        except Exception:
            weeks_count = 1

        # Determinar semana actual (índice base 0)
        cw = template_data.get('current_week') or template_data.get('Semana_num') or template_data.get('semana_num') or 1
        try:
            cw = int(cw)
        except Exception:
            cw = 1
        if cw < 1:
            cw = 1
        if cw > weeks_count:
            cw = weeks_count
        current_week_idx = cw - 1

        rx_ejercicio = re.compile(r"^\{\{\s*ejercicio_dia_(\d+)\s*\}\}$")
        rx_series = re.compile(r"^\{\{\s*series_dia_(\d+)(?:_S(\d+))?\s*\}\}$")
        rx_reps = re.compile(r"^\{\{\s*repeticiones_dia_(\d+)(?:_S(\d+))?\s*\}\}$")
        # Patrones con espacios explícitos tras '{{'
        rx_series_sp = re.compile(r"^\{\{\s+series_dia_(\d+)(?:_S(\d+))?\s*\}\}$")
        rx_reps_sp = re.compile(r"^\{\{\s+repeticiones_dia_(\d+)(?:_S(\d+))?\s*\}\}$")
        # Patrones para reemplazos incrustados en cadenas (no solo coincidencia exacta)
        rx_ejercicio_any = re.compile(r"\{\{\s*ejercicio_dia_(\d+)\s*\}\}")
        rx_series_any = re.compile(r"\{\{\s*series_dia_(\d+)(?:_S(\d+))?\s*\}\}")
        rx_reps_any = re.compile(r"\{\{\s*repeticiones_dia_(\d+)(?:_S(\d+))?\s*\}\}")
        # Variantes con espacios tras '{{'
        rx_ejercicio_any_sp = re.compile(r"\{\{\s+ejercicio_dia_(\d+)\s*\}\}")
        rx_series_any_sp = re.compile(r"\{\{\s+series_dia_(\d+)(?:_S(\d+))?\s*\}\}")
        rx_reps_any_sp = re.compile(r"\{\{\s+repeticiones_dia_(\d+)(?:_S(\d+))?\s*\}\}")
        
        # Patrones para 'semana' contextuales (con o sin mayúscula y espacios)
        rx_semana_any = re.compile(r"\{\{\s*(?:semana|Semana)\s*\}\}")
        # Pre-escanear columnas que contienen '{{ semana }}' para asignar números por posición izquierda->derecha
        semana_columns: List[int] = []
        try:
            from openpyxl.utils import column_index_from_string
        except Exception:
            column_index_from_string = None  # type: ignore
        for row_scan in sheet.iter_rows():
            for cell_scan in row_scan:
                v = cell_scan.value
                if isinstance(v, str) and rx_semana_any.search(v):
                    try:
                        col_idx = getattr(cell_scan, 'col_idx', None)
                        if col_idx is None:
                            c = getattr(cell_scan, 'column', None)
                            if isinstance(c, int):
                                col_idx = c
                            elif isinstance(c, str) and column_index_from_string:
                                col_idx = column_index_from_string(c)
                        if isinstance(col_idx, int) and col_idx not in semana_columns:
                            semana_columns.append(col_idx)
                    except Exception:
                        pass
        semana_columns.sort()
        semana_col_to_num: Dict[int, int] = {c: i + 1 for i, c in enumerate(semana_columns)}
        
        # Recorrer todas las celdas y reemplazar variables
        for row in sheet.iter_rows():
            # Llevar registro si en esta fila avanzamos el puntero de algún día
            incrementar_dias: Dict[int, bool] = {}
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # Verificar si es una celda combinada
                    if hasattr(cell, '__class__') and cell.__class__.__name__ == 'MergedCell':
                        self.logger.warning(f"Celda combinada encontrada en {cell.coordinate}, saltando")
                        continue
                    
                    original_value = cell.value
                    stripped = original_value.strip()

                    # Reemplazo contextual de '{{ semana }}' por columna (si corresponde)
                    if rx_semana_any.search(original_value):
                        try:
                            col_idx = getattr(cell, 'col_idx', None)
                            if col_idx is None:
                                c = getattr(cell, 'column', None)
                                if isinstance(c, int):
                                    col_idx = c
                                elif isinstance(c, str) and column_index_from_string:
                                    col_idx = column_index_from_string(c)
                            week_num_ctx = semana_col_to_num.get(col_idx)
                            if isinstance(week_num_ctx, int):
                                contextual = f"Semana {week_num_ctx}"
                                new_text = rx_semana_any.sub(contextual, original_value)
                                if new_text != cell.value:
                                    try:
                                        cell.value = new_text
                                    except Exception as e:
                                        self.logger.warning(f"No se pudo actualizar celda {cell.coordinate}: {e}")
                                original_value = cell.value
                                stripped = original_value.strip()
                        except Exception:
                            pass
                    
                    # 1) Distribución por filas si la celda es EXACTAMENTE el placeholder
                    m_e = rx_ejercicio.match(stripped)
                    if m_e:
                        day = int(m_e.group(1))
                        ejercicios = ejercicios_por_dia.get(day, [])
                        idx = puntero_por_dia.get(day, 0)
                        nombre = ejercicios[idx]['nombre'] if idx < len(ejercicios) else ''
                        try:
                            cell.value = nombre
                        except Exception as e:
                            self.logger.warning(f"No se pudo actualizar celda {cell.coordinate}: {e}")
                        incrementar_dias[day] = True
                        continue
                    
                    m_s = rx_series.match(stripped) or rx_series_sp.match(stripped)
                    if m_s:
                        day = int(m_s.group(1))
                        semana_idx = int(m_s.group(2)) - 1 if m_s.group(2) else current_week_idx
                        ejercicios = ejercicios_por_dia.get(day, [])
                        idx = puntero_por_dia.get(day, 0)
                        if idx < len(ejercicios):
                            series_weeks = ejercicios[idx].get('series_semanas') or []
                            valor = series_weeks[semana_idx] if semana_idx < len(series_weeks) else ''
                        else:
                            valor = ''
                        try:
                            cell.value = valor
                        except Exception as e:
                            self.logger.warning(f"No se pudo actualizar celda {cell.coordinate}: {e}")
                        continue
                    
                    m_r = rx_reps.match(stripped) or rx_reps_sp.match(stripped)
                    if m_r:
                        day = int(m_r.group(1))
                        semana_idx = int(m_r.group(2)) - 1 if m_r.group(2) else current_week_idx
                        ejercicios = ejercicios_por_dia.get(day, [])
                        idx = puntero_por_dia.get(day, 0)
                        if idx < len(ejercicios):
                            reps_weeks = ejercicios[idx].get('repeticiones_semanas') or []
                            valor = reps_weeks[semana_idx] if semana_idx < len(reps_weeks) else ''
                        else:
                            valor = ''
                        try:
                            cell.value = valor
                        except Exception as e:
                            self.logger.warning(f"No se pudo actualizar celda {cell.coordinate}: {e}")
                        continue
                    
                    # 1b) Reemplazos incrustados en cadenas para ejercicio/series/repeticiones y variantes con espacios
                    new_value = original_value
                    replaced_embedded = False
                    if rx_ejercicio_any.search(new_value):
                        def _repl_ej(m):
                            day = int(m.group(1))
                            ejercicios = ejercicios_por_dia.get(day, [])
                            idx = puntero_por_dia.get(day, 0)
                            return ejercicios[idx]['nombre'] if idx < len(ejercicios) else ''
                        new_value = rx_ejercicio_any.sub(_repl_ej, new_value)
                        replaced_embedded = True
                    if rx_series_any.search(new_value):
                        def _repl_series(m):
                            day = int(m.group(1))
                            semana_idx = int(m.group(2)) - 1 if m.group(2) else current_week_idx
                            ejercicios = ejercicios_por_dia.get(day, [])
                            idx = puntero_por_dia.get(day, 0)
                            if idx < len(ejercicios):
                                sw = ejercicios[idx].get('series_semanas') or []
                                return sw[semana_idx] if semana_idx < len(sw) else ''
                            return ''
                        new_value = rx_series_any.sub(_repl_series, new_value)
                        replaced_embedded = True
                    if rx_reps_any.search(new_value):
                        def _repl_reps(m):
                            day = int(m.group(1))
                            semana_idx = int(m.group(2)) - 1 if m.group(2) else current_week_idx
                            ejercicios = ejercicios_por_dia.get(day, [])
                            idx = puntero_por_dia.get(day, 0)
                            if idx < len(ejercicios):
                                rw = ejercicios[idx].get('repeticiones_semanas') or []
                                return rw[semana_idx] if semana_idx < len(rw) else ''
                            return ''
                        new_value = rx_reps_any.sub(_repl_reps, new_value)
                        replaced_embedded = True
                    # variantes con espacios
                    if rx_ejercicio_any_sp.search(new_value):
                        new_value = rx_ejercicio_any_sp.sub(_repl_ej, new_value)
                        replaced_embedded = True
                    if rx_series_any_sp.search(new_value):
                        new_value = rx_series_any_sp.sub(_repl_series, new_value)
                        replaced_embedded = True
                    if rx_reps_any_sp.search(new_value):
                        new_value = rx_reps_any_sp.sub(_repl_reps, new_value)
                        replaced_embedded = True
                    
                    # 2) Si no hubo reemplazos por día incrustados, hacer reemplazos globales de texto
                    if not replaced_embedded:
                        new_value2 = original_value
                        for variable, replacement in variable_mapping.items():
                            if variable in new_value2:
                                new_value2 = new_value2.replace(variable, str(replacement))
                        if new_value2 != cell.value:
                            try:
                                cell.value = new_value2
                                self.logger.debug(f"Reemplazado en {cell.coordinate}: '{cell.value}' -> '{new_value2}'")
                            except Exception as e:
                                self.logger.warning(f"No se pudo actualizar celda {cell.coordinate}: {e}")
                    else:
                        # Si hubo reemplazos incrustados por día, también aplicar reemplazos globales restantes
                        final_value = new_value
                        for variable, replacement in variable_mapping.items():
                            if variable in final_value:
                                final_value = final_value.replace(variable, str(replacement))
                        if final_value != cell.value:
                            try:
                                cell.value = final_value
                                self.logger.debug(f"Reemplazado en {cell.coordinate}: '{cell.value}' -> '{final_value}'")
                            except Exception as e:
                                self.logger.warning(f"No se pudo actualizar celda {cell.coordinate}: {e}")
            
            # Avanzar puntero por día una vez por fila si hubo un 'ejercicio' en esa fila
            for d in incrementar_dias.keys():
                puntero_por_dia[d] = puntero_por_dia.get(d, 0) + 1

    def _insert_logo_in_cell(self, ws, cell, logo_path: str, max_height_px: Optional[int] = None, max_width_px: Optional[int] = None) -> None:
        """
        Inserta el logo centrado en el área de destino (celda o rango combinado si aplica),
        manteniendo la proporción y ajustándolo para que quepa dentro del área disponible.
        Se prioriza la altura sobre la anchura.
        """
        try:
            # Determinar el índice de columna de la celda (compatibilidad distintas versiones)
            col_idx = getattr(cell, 'col_idx', None)
            if col_idx is None:
                c = getattr(cell, 'column', None)
                if isinstance(c, int):
                    col_idx = c
                elif isinstance(c, str):
                    try:
                        col_idx = column_index_from_string(c)
                    except Exception:
                        col_idx = None
            if not isinstance(col_idx, int):
                col_idx = cell.column if isinstance(cell.column, int) else 1

            # Rango destino (celda o rango combinado que la incluye)
            target_min_row = cell.row
            target_max_row = cell.row
            target_min_col = col_idx
            target_max_col = col_idx

            try:
                for rng in getattr(ws.merged_cells, 'ranges', []) or []:
                    # Comprobar si la celda está dentro del merge por límites
                    if (rng.min_row <= cell.row <= rng.max_row) and (rng.min_col <= col_idx <= rng.max_col):
                        target_min_row = rng.min_row
                        target_max_row = rng.max_row
                        target_min_col = rng.min_col
                        target_max_col = rng.max_col
                        break
            except Exception:
                pass

            # Conversión aproximada a píxeles: ancho de columna (~7 px por unidad), altura de fila (pt -> px ~1.33)
            def col_width_to_px(col_letter: str) -> float:
                try:
                    w = ws.column_dimensions[col_letter].width
                    if w is None:
                        w = 8.43  # ancho por defecto Excel
                except Exception:
                    w = 8.43
                return float(w) * 7.0

            def row_height_to_px(r: int) -> float:
                try:
                    h = ws.row_dimensions[r].height
                    if h is None:
                        return 20.0  # ~ 15pt * 1.33
                    return float(h) * 1.33
                except Exception:
                    return 20.0

            total_width_px = 0.0
            for c_idx in range(target_min_col, target_max_col + 1):
                total_width_px += col_width_to_px(get_column_letter(c_idx))
            total_height_px = 0.0
            for r_idx in range(target_min_row, target_max_row + 1):
                total_height_px += row_height_to_px(r_idx)

            # Limitar por máximos externos si se proporcionan
            if isinstance(max_width_px, (int, float)):
                total_width_px = min(total_width_px, float(max_width_px))
            if isinstance(max_height_px, (int, float)):
                total_height_px = min(total_height_px, float(max_height_px))

            # Cargar imagen y obtener tamaño original
            img = XLImage(logo_path)
            orig_w = float(getattr(img, 'width', 1)) or 1.0
            orig_h = float(getattr(img, 'height', 1)) or 1.0

            # Escalado manteniendo aspecto: llenar altura hasta donde se pueda y quepa en ancho
            scale_h = total_height_px / orig_h if orig_h > 0 else 1.0
            scale_w = total_width_px / orig_w if orig_w > 0 else 1.0
            scale = min(scale_h, scale_w) if (scale_h > 0 and scale_w > 0) else 1.0
            new_w = max(1.0, orig_w * scale)
            new_h = max(1.0, orig_h * scale)
            try:
                img.width = int(round(new_w))
                img.height = int(round(new_h))
            except Exception:
                pass

            # Centrar: calcular offset relativo dentro del área destino
            offset_x_px = max(0.0, (total_width_px - new_w) / 2.0)
            offset_y_px = max(0.0, (total_height_px - new_h) / 2.0)
            anchor_cell = ws.cell(row=target_min_row, column=target_min_col)
            anchor_coord = anchor_cell.coordinate

            try:
                if AnchorMarker and OneCellAnchor and XDRPositiveSize2D:
                    def _px_to_emu(px: float) -> int:
                        return int(round(px * 9525))
                    marker = AnchorMarker(col=target_min_col - 1, colOff=_px_to_emu(offset_x_px),
                                          row=target_min_row - 1, rowOff=_px_to_emu(offset_y_px))
                    ext = XDRPositiveSize2D(cx=_px_to_emu(new_w), cy=_px_to_emu(new_h))
                    img.anchor = OneCellAnchor(_from=marker, ext=ext)
                    ws.add_image(img)
                else:
                    ws.add_image(img, anchor_coord)
            except Exception:
                ws.add_image(img, anchor_coord)
        except Exception as e:
            try:
                self.logger.warning(f"No se pudo insertar el logo en {cell.coordinate}: {e}")
            except Exception:
                pass

    def _load_exercises_sequential(self, sheet, dias_data: List[Dict[str, Any]], current_week: Optional[int] = None):
        """
        Rellena una tabla anterior sin placeholders Jinja buscando encabezados comunes
        (Ejercicio, Series, Repeticiones) y completando fila por fila.
        
        Args:
            sheet: Hoja de Excel
            dias_data: Datos de los días con ejercicios
            current_week: Semana actual para elegir valores base de series/reps
        """
        try:
            # Detectar fila de encabezados
            header_row_idx = None
            col_ejercicio = None
            col_series = None
            col_reps = None
            series_week_cols: Dict[int, int] = {}   # semana -> col
            reps_week_cols: Dict[int, int] = {}     # semana -> col
            max_rows_scan = min(sheet.max_row, 200)
            for r in range(1, max_rows_scan + 1):
                row_vals = []
                for c in range(1, sheet.max_column + 1):
                    v = sheet.cell(row=r, column=c).value
                    row_vals.append(v)
                row_text = [str(v).strip() for v in row_vals if isinstance(v, str) and v.strip()]
                lower_map = {i+1: row_vals[i].strip().lower() if isinstance(row_vals[i], str) else '' for i in range(len(row_vals))}
                if any('ejercicio' in t for t in (v.lower() for v in row_text)):
                    header_row_idx = r
                    # Identificar columnas por nombre
                    for col_idx, text in lower_map.items():
                        if not text:
                            continue
                        if 'ejercicio' in text and col_ejercicio is None:
                            col_ejercicio = col_idx
                        if 'serie' in text and col_series is None and not re.search(r'\br\d+\b', text):
                            # Evitar confundir con R1, R2
                            col_series = col_idx
                        if any(x in text for x in ['repet', 'repes']) and col_reps is None:
                            col_reps = col_idx
                        # Series por semana: S1, S2, Series S1
                        m_s = re.match(r'^(?:series\s*)?s(\d+)$', text)
                        if m_s:
                            semana = int(m_s.group(1))
                            series_week_cols[semana] = col_idx
                        # Reps por semana: R1, Reps S1, Repeticiones S1 o R1
                        m_r = re.match(r'^(?:reps?|repeticiones?\s*)s?(\d+)$', text)
                        if m_r:
                            semana = int(m_r.group(1))
                            reps_week_cols[semana] = col_idx
                        else:
                            m_r_short = re.match(r'^r(\d+)$', text)
                            if m_r_short:
                                semana = int(m_r_short.group(1))
                                reps_week_cols[semana] = col_idx
                    break
            if header_row_idx is None or col_ejercicio is None:
                self.logger.info("No se encontraron encabezados de tabla anteriores (Ejercicio/Series/Repeticiones). Omitiendo secuencial.")
                return
            start_row = header_row_idx + 1
            write_row = start_row

            # Asegurar que no intentamos escribir dentro de celdas fusionadas (MergedCell)
            def _adjust_to_unmerged_row(row: int) -> int:
                target_cols = set()
                if col_ejercicio:
                    target_cols.add(col_ejercicio)
                if col_series:
                    target_cols.add(col_series)
                if col_reps:
                    target_cols.add(col_reps)
                for _c in series_week_cols.values():
                    target_cols.add(_c)
                for _c in reps_week_cols.values():
                    target_cols.add(_c)
                while True:
                    adjusted = row
                    for _c in target_cols:
                        cell = sheet.cell(row=row, column=_c)
                        if isinstance(cell, MergedCell):
                            # Encontrar el rango fusionado que cubre esta celda
                            for rng in sheet.merged_cells.ranges:
                                min_col, min_row, max_col, max_row = rng.bounds
                                if min_row <= row <= max_row and min_col <= _c <= max_col:
                                    adjusted = max(adjusted, max_row + 1)
                                    break
                    if adjusted == row:
                        return row
                    row = adjusted

            # Escribir ejercicios secuencialmente
            days_sorted = sorted(dias_data, key=lambda d: d.get('numero') or d.get('dia') or 0)
            for di, day in enumerate(days_sorted):
                ejercicios = day.get('ejercicios', []) or []
                for ej in ejercicios:
                    nombre = ej.get('nombre', '')
                    # Valores semanales
                    s_weeks = ej.get('series_semanas') or []
                    r_weeks = ej.get('repeticiones_semanas') or []
                    # Derivar valores base a partir de la semana actual si está disponible
                    if current_week is not None and isinstance(current_week, int) and current_week > 0:
                        idx_curr = current_week - 1
                        series_val = s_weeks[idx_curr] if idx_curr < len(s_weeks) else ''
                        reps_val = r_weeks[idx_curr] if idx_curr < len(r_weeks) else ''
                    else:
                        # Si no hay semana actual, no rellenar desde valores planos; dejar vacío
                        series_val = ''
                        reps_val = ''
                    try:
                        # Mover write_row a la siguiente fila libre no fusionada en las columnas objetivo
                        write_row = _adjust_to_unmerged_row(write_row)
                        sheet.cell(row=write_row, column=col_ejercicio).value = nombre
                        if col_series:
                            sheet.cell(row=write_row, column=col_series).value = series_val
                        if col_reps:
                            sheet.cell(row=write_row, column=col_reps).value = reps_val
                        # Series por semana si existen columnas S1..Sn
                        if s_weeks:
                            for semana, col_idx in series_week_cols.items():
                                idx = semana - 1
                                val = s_weeks[idx] if idx < len(s_weeks) else ''
                                cell = sheet.cell(row=write_row, column=col_idx)
                                if isinstance(cell, MergedCell):
                                    # Saltar escritura en celdas fusionadas específicas de semana
                                    continue
                                sheet.cell(row=write_row, column=col_idx).value = val
                        # Repeticiones por semana si existen columnas R1..Rn o similares
                        r_weeks = ej.get('repeticiones_semanas') or []
                        if r_weeks:
                            for semana, col_idx in reps_week_cols.items():
                                idx = semana - 1
                                val = r_weeks[idx] if idx < len(r_weeks) else ''
                                cell = sheet.cell(row=write_row, column=col_idx)
                                if isinstance(cell, MergedCell):
                                    # Saltar escritura en celdas fusionadas específicas de semana
                                    continue
                                sheet.cell(row=write_row, column=col_idx).value = val
                    except Exception as e:
                        self.logger.warning(f"No se pudo escribir ejercicio en fila {write_row}: {e}")
                    write_row += 1
                # Rellenar hasta 8 filas por día con filas en blanco para respetar el bloque por día
                try:
                    filled = len(ejercicios)
                except Exception:
                    filled = 0
                if filled < 8:
                    to_pad = 8 - filled
                    for _ in range(to_pad):
                        write_row = _adjust_to_unmerged_row(write_row)
                        # dejar fila en blanco
                        write_row += 1
                # Fila separadora en blanco entre días (exactamente 1 fila)
                if di < len(days_sorted) - 1:
                    write_row = _adjust_to_unmerged_row(write_row)
                    write_row += 1
            self.logger.info(f"Tabla anterior completada desde fila {start_row} hasta {write_row-1}")
        except Exception as e:
            self.logger.exception(f"Error en _load_exercises_sequential: {e}")
            # No relanzar para no romper exportación cuando este método no aplica
            return

    def generate_routine_pdf(self, rutina: Rutina, usuario: Usuario,
                           exercises_by_day: Dict[int, List[RutinaEjercicio]], 
                           output_path: str = None) -> str:
        """
        Genera un archivo PDF de rutina.
        
        Args:
            rutina: Objeto Rutina
            usuario: Objeto Usuario
            exercises_by_day: Diccionario con ejercicios por día
            output_path: Ruta de salida personalizada
            
        Returns:
            Ruta del archivo generado
        """
        try:
            self.logger.info("Iniciando generación de PDF de rutina")
            
            # Validar datos
            is_valid, errors = self.validate_routine_data(rutina, usuario, exercises_by_day)
            if not is_valid:
                error_msg = f"Datos de rutina inválidos: {', '.join(errors)}"
                self.logger.error(error_msg)
                raise ValueError(error_msg)
            
            # Generar nombre de archivo si no se proporciona
            if not output_path:
                fecha_str = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"rutina_{usuario.nombre.replace(' ', '_')}_{fecha_str}.pdf"
                output_path = self.output_dir_pdf / filename
            else:
                output_path = Path(output_path)
            
            # Crear documento PDF en orientación horizontal
            doc = SimpleDocTemplate(str(output_path), pagesize=landscape(A4))
            styles = getSampleStyleSheet()
            story = []
            
            # Título
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=TA_CENTER
            )
            story.append(Paragraph("ZURKA FITNESS - RUTINA PERSONALIZADA", title_style))
            
            # Información del usuario
            info_style = styles['Normal']
            story.append(Paragraph(f"<b>Cliente:</b> {usuario.nombre}", info_style))
            story.append(Paragraph(f"<b>Rutina:</b> {rutina.nombre_rutina}", info_style))
            story.append(Paragraph(f"<b>Fecha:</b> {datetime.now().strftime('%d/%m/%Y')}", info_style))
            story.append(Spacer(1, 20))
            
            # Ejercicios por día
            for day_num in sorted(exercises_by_day.keys()):
                exercises = exercises_by_day[day_num]
                
                # Título del día
                day_title = Paragraph(f"<b>DÍA {day_num}</b>", styles['Heading2'])
                story.append(day_title)
                
                # Tabla de ejercicios
                table_data = [['Ejercicio', 'Series', 'Repeticiones']]
                
                for exercise in exercises:
                    # Igual que en Excel, soportar atributo personalizado
                    nombre_e = getattr(exercise, 'nombre_ejercicio', None) or (exercise.ejercicio.nombre if getattr(exercise, 'ejercicio', None) else 'Sin nombre')
                    table_data.append([
                        nombre_e,
                        getattr(exercise, 'series', '') or '',
                        getattr(exercise, 'repeticiones', '') or ''
                    ])
                
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 14),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(table)
                story.append(Spacer(1, 20))
            
            # Construir PDF
            doc.build(story)
            
            self.logger.info(f"Rutina PDF generada exitosamente: {output_path}")
            return str(output_path)
            
        except Exception as e:
            self.logger.error(f"Error al generar PDF de rutina: {e}")
            raise e

# Mantener compatibilidad con código existente
def create_routine_manager(database_manager: DatabaseManager = None) -> RoutineTemplateManager:
    """
    Factory function para crear una instancia del gestor de rutinas.
    
    Args:
        database_manager: Instancia del gestor de base de datos
        
    Returns:
        Instancia de RoutineTemplateManager
    """
    return RoutineTemplateManager(database_manager=database_manager)


def _apply_column_widths_all_sheets(self, workbook, start_col_letter: str = 'L', end_col_letter: str = 'AA', width: float = 3.71, pixel_width: Optional[int] = None) -> None:
    """Aplica un ancho fijo a columnas desde start_col_letter hasta end_col_letter en todas las hojas
    y además fuerza alineaciones de tabla (Ejercicio/Series/Repeticiones) según requerimientos.

    Args:
        workbook: Libro de trabajo openpyxl ya cargado
        start_col_letter: Letra de columna inicial (p.ej. 'A')
        end_col_letter: Letra de columna final (p.ej. 'K')
        width: Ancho en unidades de Excel (~ caracteres del estilo Normal). Se usa si no se especifica pixel_width.
        pixel_width: Ancho deseado en píxeles (entero). Si se especifica, tiene prioridad y se convierte internamente a unidades de Excel.
    """
    try:
        from openpyxl.utils import column_index_from_string, get_column_letter
        from openpyxl.cell.cell import MergedCell as _MergedCell
        # 1) Aplicar anchos
        start_idx = column_index_from_string(start_col_letter)
        end_idx = column_index_from_string(end_col_letter)
        for sh in getattr(workbook, 'worksheets', []):
            # Reajustar configuración de hoja para evitar autoajuste global
            try:
                fmt = getattr(sh, 'sheet_format', None)
                if fmt is not None:
                    try:
                        fmt.defaultColWidth = 8.43
                        fmt.baseColWidth = 10
                    except Exception:
                        pass
            except Exception:
                pass
            # Apagar bestFit en columnas existentes y neutralizar rangos que se solapen
            try:
                dims = list(getattr(sh, 'column_dimensions', {}).values())
                for _dim in dims:
                    try:
                        # Desactivar autoajuste en cualquier dimensión
                        _dim.bestFit = False
                        # Si es un rango (min..max) que interseca con nuestro objetivo, eliminar su ancho para no sobrescribir
                        dmin = getattr(_dim, 'min', None)
                        dmax = getattr(_dim, 'max', None)
                        if isinstance(dmin, int) and isinstance(dmax, int):
                            # Intersección con [start_idx, end_idx]
                            if not (dmax < start_idx or dmin > end_idx):
                                _dim.width = None
                                # Evitar que quede marcado como customWidth
                                try:
                                    _dim.customWidth = False
                                except Exception:
                                    pass
                        else:
                            # Mantener customWidth cuando ya exista un width establecido a nivel de columna
                            _dim.customWidth = bool(getattr(_dim, 'width', None) is not None)
                    except Exception:
                        continue
            except Exception:
                pass
            for col_idx in range(start_idx, end_idx + 1):
                try:
                    letter = get_column_letter(col_idx)
                    # Forzar ancho personalizado para que Excel no lo normalice y ajustarlo a incrementos reales de Excel
                    col_dim = sh.column_dimensions[letter]
                    try:
                        import math as _math
                        if pixel_width is not None:
                            # Ajuste para que el tooltip de Excel muestre exactamente los píxeles pedidos
                            # (en muchos entornos el tooltip usa px ≈ floor(7 * width))
                            target_px = max(1, int(pixel_width))
                            snapped_units = max(0.0, (target_px / 7.0) + 1e-6)
                        else:
                            target = float(width)
                            # Mantener conversión estable desde unidades de Excel
                            px = max(1, int(_math.ceil(7.0 * target + 5.0 - 1e-9)))
                            snapped_units = max(0.0, (px - 5.0) / 7.0)
                    except Exception:
                        snapped_units = float(width)
                    col_dim.width = float(snapped_units)
                    try:
                        # Reforzar para evitar autoajuste
                        col_dim.bestFit = False
                        col_dim.customWidth = True
                        try:
                            # Algunas versiones exponen auto_size
                            setattr(col_dim, 'auto_size', False)
                        except Exception:
                            pass
                    except Exception:
                        pass
                except Exception:
                    # Continuar con otras columnas/hojas incluso si una falla
                    continue
        # 2) Centrar todas las celdas en el rango de columnas solicitado (L..AA por defecto)
        from openpyxl.styles import Alignment as _Alignment
        # Solo centrar todas las celdas cuando el rango solicitado sea exactamente L..AA
        if (str(start_col_letter).upper() == 'L' and str(end_col_letter).upper() == 'AA'):
            for sh in getattr(workbook, 'worksheets', []):
                try:
                    max_rows = getattr(sh, 'max_row', 0) or 0
                    if max_rows > 0:
                        for col_idx in range(start_idx, end_idx + 1):
                            for r in range(1, max_rows + 1):
                                try:
                                    cell = sh.cell(row=r, column=col_idx)
                                    if not isinstance(cell, _MergedCell):
                                        cell.alignment = _Alignment(horizontal='center', vertical='center')
                                except Exception:
                                    continue
                except Exception:
                    continue
        # 3) Forzar alineaciones de tabla (si detectamos encabezados)
        import re as _re
        for sh in getattr(workbook, 'worksheets', []):
            try:
                header_row_idx = None
                col_ejercicio = None
                col_series = None
                col_reps = None
                series_week_cols = {}
                reps_week_cols = {}
                max_rows_scan = min(getattr(sh, 'max_row', 0) or 0, 200)
                if max_rows_scan <= 0:
                    continue
                for r in range(1, max_rows_scan + 1):
                    row_vals = [sh.cell(row=r, column=c).value for c in range(1, sh.max_column + 1)]
                    texts_lower = {i+1: (str(v).strip().lower() if isinstance(v, str) else '') for i, v in enumerate(row_vals)}
                    if any(('ejercicio' in t) for t in texts_lower.values() if t):
                        header_row_idx = r
                        for col_idx, text in texts_lower.items():
                            if not text:
                                continue
                            if ('ejercicio' in text) and col_ejercicio is None:
                                col_ejercicio = col_idx
                            if ('serie' in text) and (col_series is None) and (not _re.search(r'\br\d+\b', text)):
                                col_series = col_idx
                            if (('repet' in text) or ('repes' in text)) and col_reps is None:
                                col_reps = col_idx
                            m_s = _re.match(r'^(?:series\s*)?s(\d+)$', text)
                            if m_s:
                                semana = int(m_s.group(1))
                                series_week_cols[semana] = col_idx
                            m_r = _re.match(r'^(?:reps?|repeticiones?\s*)s?(\d+)$', text)
                            if m_r:
                                semana = int(m_r.group(1))
                                reps_week_cols[semana] = col_idx
                            else:
                                m_r_short = _re.match(r'^r(\d+)$', text)
                                if m_r_short:
                                    semana = int(m_r_short.group(1))
                                    reps_week_cols[semana] = col_idx
                        break
                if header_row_idx is None:
                    continue
                start_row = header_row_idx + 1
                last_row = getattr(sh, 'max_row', 0) or 0
                if last_row <= start_row:
                    continue
                # Alineaciones según requerimiento
                for r in range(start_row, last_row + 1):
                    try:
                        if col_ejercicio:
                            cell = sh.cell(row=r, column=col_ejercicio)
                            if not isinstance(cell, _MergedCell):
                                cell.alignment = _Alignment(horizontal='left', vertical='center', wrap_text=True)
                        if col_series:
                            cell = sh.cell(row=r, column=col_series)
                            if not isinstance(cell, _MergedCell):
                                cell.alignment = _Alignment(horizontal='center', vertical='center')
                        if col_reps:
                            cell = sh.cell(row=r, column=col_reps)
                            if not isinstance(cell, _MergedCell):
                                cell.alignment = _Alignment(horizontal='center', vertical='center')
                        for _c in series_week_cols.values():
                            cell = sh.cell(row=r, column=_c)
                            if not isinstance(cell, _MergedCell):
                                cell.alignment = _Alignment(horizontal='center', vertical='center')
                        for _c in reps_week_cols.values():
                            cell = sh.cell(row=r, column=_c)
                            if not isinstance(cell, _MergedCell):
                                cell.alignment = _Alignment(horizontal='center', vertical='center')
                    except Exception:
                        continue
            except Exception:
                continue
    except Exception as e:
        # No romper generación si no se pueden aplicar anchos/alineaciones
        try:
            self.logger.warning(f"No se pudieron aplicar anchos/alineaciones de columna ({start_col_letter}:{end_col_letter}): {e}")
        except Exception:
            pass